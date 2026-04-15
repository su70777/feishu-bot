from __future__ import annotations

import base64
import csv
import json
import hmac
import hashlib
import io
import secrets
import subprocess
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
import time
from typing import Any

import requests
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from pydantic import BaseModel, Field


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
TEMPLATE_FILE = DATA_DIR / "recipient_templates.json"
SEND_LOG_FILE = DATA_DIR / "send_logs.json"
PROFILE_FILE = DATA_DIR / "student_profiles.json"
DEBUG_LOG_FILE = DATA_DIR / "event_debug.jsonl"
ROSTER_FILE = DATA_DIR / "student_openid_map.json"
EVENT_CACHE_FILE = DATA_DIR / "processed_event_ids.json"
RECENT_COMMAND_CACHE_FILE = DATA_DIR / "recent_commands.json"
CHAT_TAG_FILE = DATA_DIR / "chat_tags.json"
LAST_SEND_CTX_FILE = DATA_DIR / "last_send_context.json"
SEND_RECORD_FILE = DATA_DIR / "send_records.json"
ORG_CONTACT_FILE = DATA_DIR / "org_contacts.json"
STUDENT_TAG_FILE = DATA_DIR / "student_tags.json"
AUTHORIZED_USER_FILE = DATA_DIR / "authorized_users.json"
ALERT_RECEIVER_FILE = DATA_DIR / "alert_receivers.json"
LAST_UPLOADED_NAME_FILE = DATA_DIR / "last_uploaded_name_files.json"
AUTO_REPLY_SETTING_FILE = DATA_DIR / "auto_reply_settings.json"
AUTO_REPLY_FAQ_FILE = DATA_DIR / "auto_reply_faq.json"
AUTO_REPLY_STATE_FILE = DATA_DIR / "auto_reply_state.json"
USER_TOKEN_FILE = DATA_DIR / "user_tokens.json"

MANAGE_SESSION_COOKIE = "feishu_manage_session"
MANAGE_SERVICE_NAME = "feishu-bot"
MANAGE_SESSION_MAX_AGE = 7 * 24 * 3600

FEISHU_BASE = "https://open.feishu.cn/open-apis"
FEISHU_AUTH_BASE = "https://accounts.feishu.cn/open-apis/authen/v1"
BOT_BUILD = "2026-04-15.4"


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        # 容错：JSON 文件损坏时不让主流程中断，避免机器人“发了没反应”。
        return default


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _append_debug(payload: dict[str, Any]) -> None:
    row = {"ts": datetime.now().isoformat(), **payload}
    DEBUG_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with DEBUG_LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")


def _norm_text(v: str) -> str:
    return " ".join((v or "").strip().split())


def _strip_bot_mention_prefix(text: str) -> str:
    value = (text or "").strip()
    if not value:
        return ""
    while value.startswith("@"):
        parts = value.split(maxsplit=1)
        if len(parts) == 1:
            return ""
        value = parts[1].strip()
    return value


def _split_csv_names(v: str) -> list[str]:
    if not v:
        return []
    raw = v.replace("，", ",").replace("、", ",")
    return [x.strip() for x in raw.split(",") if x.strip()]


def _parse_message_content(raw_content: Any) -> dict[str, Any]:
    if isinstance(raw_content, dict):
        return raw_content
    if isinstance(raw_content, str) and raw_content.strip():
        try:
            value = json.loads(raw_content)
            return value if isinstance(value, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _load_last_uploaded_name_rows() -> dict[str, Any]:
    rows = _read_json(LAST_UPLOADED_NAME_FILE, {})
    return rows if isinstance(rows, dict) else {}


def _save_last_uploaded_name_rows(rows: dict[str, Any]) -> None:
    _write_json(LAST_UPLOADED_NAME_FILE, rows if isinstance(rows, dict) else {})


def _save_last_uploaded_name_file(
    sender_open_id: str,
    *,
    chat_id: str,
    file_name: str,
    names: list[str],
    message_id: str,
    file_key: str,
) -> None:
    if not sender_open_id:
        return
    rows = _load_last_uploaded_name_rows()
    rows[sender_open_id] = {
        "chat_id": chat_id,
        "file_name": file_name,
        "names": _dedup_keep_order(names),
        "message_id": message_id,
        "file_key": file_key,
        "updated_at": datetime.now().isoformat(),
    }
    _save_last_uploaded_name_rows(rows)


def _get_last_uploaded_name_file(sender_open_id: str) -> dict[str, Any]:
    if not sender_open_id:
        return {}
    rows = _load_last_uploaded_name_rows()
    row = rows.get(sender_open_id, {})
    return row if isinstance(row, dict) else {}


def _clean_name_candidate(value: str) -> str:
    text = _norm_text(str(value or ""))
    text = re.sub(r"^[0-9]+[.、)\s-]*", "", text).strip()
    text = text.strip(" ,，;；|/\\")
    if not text:
        return ""
    lowered = text.lower()
    if lowered in {"姓名", "名字", "name", "student", "student name", "学员", "学生"}:
        return ""
    if text.startswith("说明") or text.startswith("备注"):
        return ""
    if len(text) > 40:
        return ""
    return text


def _extract_names_from_rows(rows: list[list[str]]) -> list[str]:
    normalized_rows: list[list[str]] = []
    for row in rows:
        cells = [_norm_text(str(cell or "")) for cell in row]
        if any(cells):
            normalized_rows.append(cells)
    if not normalized_rows:
        return []

    header_keywords = ("姓名", "名字", "name", "学生", "学员")
    name_col_idx = 0
    header_row_idx = -1
    for ridx, row in enumerate(normalized_rows[:5]):
        for cidx, cell in enumerate(row):
            lowered = cell.lower()
            if any(keyword in lowered for keyword in header_keywords):
                name_col_idx = cidx
                header_row_idx = ridx
                break
        if header_row_idx >= 0:
            break

    start_idx = header_row_idx + 1 if header_row_idx >= 0 else 0
    results: list[str] = []
    for row in normalized_rows[start_idx:]:
        if name_col_idx < len(row):
            candidate = _clean_name_candidate(row[name_col_idx])
            if candidate:
                results.append(candidate)
                continue
        for cell in row:
            candidate = _clean_name_candidate(cell)
            if candidate:
                results.append(candidate)
                break
    return _dedup_keep_order(results)


def _decode_bytes_to_text(blob: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030", "utf-16"):
        try:
            return blob.decode(encoding)
        except UnicodeDecodeError:
            continue
    return blob.decode("utf-8", errors="ignore")


def _extract_names_from_text_blob(text: str) -> list[str]:
    lines = text.replace("\r", "\n").split("\n")
    rows: list[list[str]] = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if any(sep in stripped for sep in [",", "，", "\t", ";", "；", "|", "、"]):
            parts = [part.strip() for part in re.split(r"[,，\t;；|、]+", stripped) if part.strip()]
            rows.append(parts)
        else:
            rows.append([stripped])
    return _extract_names_from_rows(rows)


def _extract_names_from_csv_blob(blob: bytes) -> list[str]:
    text = _decode_bytes_to_text(blob)
    reader = csv.reader(io.StringIO(text))
    return _extract_names_from_rows([[str(cell) for cell in row] for row in reader])


def _extract_names_from_xlsx_blob(blob: bytes) -> tuple[list[str], str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return [], "当前环境缺少 openpyxl，暂时无法解析 xlsx 文件。"

    try:
        workbook = load_workbook(filename=io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as exc:
        return [], f"xlsx 解析失败：{exc}"

    rows: list[list[str]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if cell is None else str(cell) for cell in row])
    return _extract_names_from_rows(rows), ""


def _parse_uploaded_name_file(file_name: str, blob: bytes) -> tuple[list[str], str]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".xlsx":
        return _extract_names_from_xlsx_blob(blob)
    if suffix in {".csv"}:
        return _extract_names_from_csv_blob(blob), ""
    if suffix in {".txt", ".md"} or not suffix:
        return _extract_names_from_text_blob(_decode_bytes_to_text(blob)), ""
    return [], "当前仅支持 txt、csv、xlsx 名单文件。"


def _download_message_resource(message_id: str, file_key: str, resource_type: str = "file") -> tuple[bytes, str]:
    if not message_id or not file_key:
        return b"", "缺少 message_id 或 file_key"
    if SETTINGS.mock_send:
        return b"", "当前为 mock 模式，无法下载飞书文件"
    token = _token()
    headers = {"Authorization": f"Bearer {token}"}
    urls = [
        f"{FEISHU_BASE}/im/v1/messages/{message_id}/resources/{file_key}?type={resource_type}",
        f"{FEISHU_BASE}/im/v1/messages/{message_id}/resources/{file_key}",
    ]
    last_error = ""
    for url in urls:
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200 and resp.content:
                return resp.content, ""
            try:
                payload = resp.json()
            except Exception:
                payload = resp.text
            last_error = f"HTTP {resp.status_code}: {payload}"
        except Exception as exc:
            last_error = str(exc)
    return b"", last_error or "下载失败"


def _replace_students_in_tag(tag_name: str, student_names: list[str]) -> list[str]:
    key = (tag_name or "").strip()
    if not key:
        return []
    tags = _load_student_tags()
    tags[key] = _dedup_keep_order([_clean_name_candidate(name) for name in student_names if _clean_name_candidate(name)])
    _save_student_tags(tags)
    return tags[key]


def _risk_score(level: str) -> int:
    if level == "高":
        return 3
    if level == "中":
        return 2
    return 1


def _is_valid_open_id(value: str) -> bool:
    if not value:
        return False
    v = value.strip()
    # 飞书 open_id 只允许字母数字下划线，且长度应明显大于示例值。
    if not re.fullmatch(r"ou_[A-Za-z0-9_]{10,}", v):
        return False
    return v.lower() not in {"ou_xxx", "ou_demo", "ou_test"}


def _consume_event_once(event_id: str) -> bool:
    if not event_id:
        return True
    ids = _read_json(EVENT_CACHE_FILE, [])
    if event_id in ids:
        return False
    ids.append(event_id)
    if len(ids) > 1000:
        ids = ids[-1000:]
    _write_json(EVENT_CACHE_FILE, ids)
    return True


def _consume_recent_command_once(sender_open_id: str, chat_id: str, text: str, window_seconds: int = 8) -> bool:
    norm_sender = (sender_open_id or "").strip()
    norm_chat = (chat_id or "").strip()
    norm_text = _norm_text(text)
    if not norm_text:
        return True

    key = f"{norm_sender}|{norm_chat}|{norm_text}"
    now_ts = time.time()
    rows = _read_json(RECENT_COMMAND_CACHE_FILE, {})
    if not isinstance(rows, dict):
        rows = {}

    valid_rows: dict[str, float] = {}
    for k, ts in rows.items():
        try:
            ts_value = float(ts)
        except (TypeError, ValueError):
            continue
        # 仅保留“过去 window_seconds 秒内”的记录；
        # 若出现未来时间戳（例如服务器时间回拨/漂移），直接丢弃，避免长期误判重复。
        if 0 <= now_ts - ts_value <= window_seconds:
            valid_rows[k] = ts_value

    last_ts = valid_rows.get(key)
    if last_ts is not None and now_ts - last_ts <= window_seconds:
        _write_json(RECENT_COMMAND_CACHE_FILE, valid_rows)
        return False

    valid_rows[key] = now_ts
    _write_json(RECENT_COMMAND_CACHE_FILE, valid_rows)
    return True


@dataclass
class Settings:
    app_id: str = ""
    app_secret: str = ""
    verify_token: str = ""
    teacher_open_id: str = ""
    org_department_id: str = "0"
    # 低/中/高。中表示中高风险才告警。
    alert_level: str = "中"
    mock_send: bool = True
    ai_base_url: str = ""
    ai_api_key: str = ""
    ai_model: str = ""
    public_base_url: str = ""
    oauth_state_secret: str = ""


SETTINGS = Settings()


def _load_dotenv() -> dict[str, str]:
    env_file = BASE_DIR / ".env"
    if not env_file.exists():
        return {}
    values: dict[str, str] = {}
    for line in env_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        values[k.strip()] = v.strip()
    return values


def _token() -> str:
    if SETTINGS.mock_send:
        return "mock-token"
    if not SETTINGS.app_id or not SETTINGS.app_secret:
        raise HTTPException(status_code=400, detail="缺少 FEISHU_APP_ID / FEISHU_APP_SECRET")
    resp = requests.post(
        f"{FEISHU_BASE}/auth/v3/tenant_access_token/internal",
        json={"app_id": SETTINGS.app_id, "app_secret": SETTINGS.app_secret},
        timeout=10,
    )
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(status_code=400, detail=f"获取 tenant_access_token 失败: {data}")
    return data["tenant_access_token"]


def _send_text(receive_id: str, text: str, receive_id_type: str = "open_id") -> dict[str, Any]:
    if SETTINGS.mock_send:
        return {"mock": True, "receive_id": receive_id, "text": text, "receive_id_type": receive_id_type}
    token = _token()
    headers = {"Authorization": f"Bearer {token}"}
    body = {
        "receive_id": receive_id,
        "msg_type": "text",
        "content": json.dumps({"text": text}, ensure_ascii=False),
    }
    resp = requests.post(
        f"{FEISHU_BASE}/im/v1/messages?receive_id_type={receive_id_type}",
        headers=headers,
        json=body,
        timeout=10,
    )
    return resp.json()


def _send_card(receive_id: str, card: dict[str, Any], receive_id_type: str = "open_id") -> dict[str, Any]:
    if SETTINGS.mock_send:
        return {"mock": True, "receive_id": receive_id, "card": card, "receive_id_type": receive_id_type}
    token = _token()
    headers = {"Authorization": f"Bearer {token}"}
    body = {
        "receive_id": receive_id,
        "msg_type": "interactive",
        "content": json.dumps(card, ensure_ascii=False),
    }
    resp = requests.post(
        f"{FEISHU_BASE}/im/v1/messages?receive_id_type={receive_id_type}",
        headers=headers,
        json=body,
        timeout=10,
    )
    return resp.json()


def _reply_chat(chat_id: str, text: str) -> dict[str, Any]:
    if not chat_id:
        return {"code": -1, "msg": "chat_id missing"}
    return _send_text(chat_id, text, receive_id_type="chat_id")


def _load_user_tokens() -> dict[str, Any]:
    rows = _read_json(USER_TOKEN_FILE, {})
    return rows if isinstance(rows, dict) else {}


def _save_user_tokens(rows: dict[str, Any]) -> None:
    _write_json(USER_TOKEN_FILE, rows if isinstance(rows, dict) else {})


def _oauth_redirect_uri() -> str:
    base = (SETTINGS.public_base_url or "").strip().rstrip("/")
    if not base:
        raise HTTPException(status_code=400, detail="缺少 PUBLIC_BASE_URL，请先在 .env 中配置固定域名")
    return f"{base}/auth/feishu/callback"


def _state_secret_bytes() -> bytes:
    return (SETTINGS.oauth_state_secret or SETTINGS.app_secret or "state-secret").encode("utf-8")


def _encode_state_blob(payload: dict[str, Any]) -> str:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    raw_bytes = raw.encode("utf-8")
    sign = hmac.new(_state_secret_bytes(), raw_bytes, hashlib.sha256).hexdigest()
    token = base64.urlsafe_b64encode(raw_bytes).decode("ascii").rstrip("=")
    return f"{token}.{sign}"


def _decode_state_blob(value: str) -> dict[str, Any] | None:
    if not value:
        return None
    secret = _state_secret_bytes()
    if "." in value:
        token, sign = value.rsplit(".", 1)
        try:
            payload_bytes = base64.urlsafe_b64decode(token + "=" * (-len(token) % 4))
            raw = payload_bytes.decode("utf-8")
        except Exception:
            return None
        expected = hmac.new(secret, raw.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sign, expected):
            return None
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            return None
        if not isinstance(payload, dict):
            return None
        try:
            ts_value = int(payload.get("ts", 0))
        except (TypeError, ValueError):
            return None
        if abs(int(time.time()) - ts_value) > 1800:
            return None
        payload["ts"] = ts_value
        payload["purpose"] = str(payload.get("purpose") or "login")
        payload["next"] = str(payload.get("next") or "")
        return payload

    # Backward-compatible fallback for the original ts:nonce:sign format.
    if value.count(":") != 2:
        return None
    ts, nonce, sign = value.split(":", 2)
    raw = f"{ts}:{nonce}"
    expected = hmac.new(secret, raw.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sign, expected):
        return None
    try:
        ts_value = int(ts)
    except ValueError:
        return None
    if abs(int(time.time()) - ts_value) > 1800:
        return None
    return {"ts": ts_value, "nonce": nonce, "purpose": "login", "next": ""}


def _build_oauth_state(*, purpose: str = "login", next_path: str = "") -> str:
    payload = {
        "ts": int(time.time()),
        "nonce": secrets.token_urlsafe(8),
        "purpose": purpose,
        "next": next_path,
    }
    return _encode_state_blob(payload)


def _verify_oauth_state(value: str) -> bool:
    return _decode_state_blob(value) is not None


def _oauth_login_url(*, purpose: str = "login", next_path: str = "") -> str:
    redirect_uri = _oauth_redirect_uri()
    state = _build_oauth_state(purpose=purpose, next_path=next_path)
    return (
        f"{FEISHU_AUTH_BASE}/authorize"
        f"?app_id={SETTINGS.app_id}"
        f"&redirect_uri={requests.utils.quote(redirect_uri, safe='')}"
        f"&scope={requests.utils.quote('im:message.p2p_msg:get_as_user im:message.group_msg:get_as_user im:message.send_as_user', safe='')}"
        f"&state={requests.utils.quote(state, safe='')}"
    )


def _manage_page_url() -> str:
    base = (SETTINGS.public_base_url or "").strip().rstrip("/")
    return f"{base}/manage" if base else ""


def _exchange_user_access_token(code: str) -> dict[str, Any]:
    if not SETTINGS.app_id or not SETTINGS.app_secret:
        raise HTTPException(status_code=400, detail="缺少 FEISHU_APP_ID / FEISHU_APP_SECRET")
    resp = requests.post(
        f"{FEISHU_BASE}/authen/v1/oidc/access_token",
        json={
            "grant_type": "authorization_code",
            "code": code,
            "app_id": SETTINGS.app_id,
            "app_secret": SETTINGS.app_secret,
            "redirect_uri": _oauth_redirect_uri(),
        },
        timeout=20,
    )
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(status_code=400, detail=f"换取 user_access_token 失败: {data}")
    return data.get("data") or {}


def _refresh_user_access_token(refresh_token: str) -> dict[str, Any]:
    resp = requests.post(
        f"{FEISHU_BASE}/authen/v1/refresh_access_token",
        json={
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "app_id": SETTINGS.app_id,
            "app_secret": SETTINGS.app_secret,
        },
        timeout=20,
    )
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(status_code=400, detail=f"刷新 user_access_token 失败: {data}")
    return data.get("data") or {}


def _fetch_user_identity(user_access_token: str) -> dict[str, Any]:
    resp = requests.get(
        f"{FEISHU_BASE}/authen/v1/user_info",
        headers={"Authorization": f"Bearer {user_access_token}"},
        timeout=15,
    )
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(status_code=400, detail=f"获取用户身份失败: {data}")
    return data.get("data") or {}


def _store_user_token_bundle(token_data: dict[str, Any], user_info: dict[str, Any]) -> dict[str, Any]:
    open_id = (user_info.get("open_id") or "").strip()
    if not open_id:
        raise HTTPException(status_code=400, detail="用户授权成功，但未获取到 open_id")
    rows = _load_user_tokens()
    rows[open_id] = {
        "name": (user_info.get("name") or "").strip(),
        "open_id": open_id,
        "union_id": (user_info.get("union_id") or "").strip(),
        "user_id": (user_info.get("user_id") or "").strip(),
        "access_token": token_data.get("access_token", ""),
        "refresh_token": token_data.get("refresh_token", ""),
        "expires_in": token_data.get("expires_in", 0),
        "refresh_expires_in": token_data.get("refresh_expires_in", 0),
        "updated_at": int(time.time()),
    }
    _save_user_tokens(rows)
    return rows[open_id]


def _get_teacher_user_token_bundle() -> dict[str, Any] | None:
    teacher_open_id = (SETTINGS.teacher_open_id or "").strip()
    if not teacher_open_id:
        return None
    rows = _load_user_tokens()
    bundle = rows.get(teacher_open_id)
    if not isinstance(bundle, dict):
        return None
    return bundle


def _build_manage_session(open_id: str) -> str:
    ts = str(int(time.time()))
    raw = f"{open_id}:{ts}"
    sign = hmac.new(_state_secret_bytes(), raw.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{raw}:{sign}"


def _parse_manage_session(value: str) -> str:
    if not value or value.count(":") < 2:
        return ""
    open_id, ts, sign = value.rsplit(":", 2)
    raw = f"{open_id}:{ts}"
    expected = hmac.new(_state_secret_bytes(), raw.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sign, expected):
        return ""
    try:
        ts_value = int(ts)
    except ValueError:
        return ""
    if abs(int(time.time()) - ts_value) > MANAGE_SESSION_MAX_AGE:
        return ""
    return open_id.strip()


def _manage_open_id_from_request(request: Request) -> str:
    open_id = _parse_manage_session(request.cookies.get(MANAGE_SESSION_COOKIE, ""))
    teacher_id = (SETTINGS.teacher_open_id or "").strip()
    if not open_id or (teacher_id and open_id != teacher_id):
        raise HTTPException(status_code=401, detail="未登录或无权访问管理页")
    return open_id


def _display_name_for_open_id(open_id: str) -> str:
    bundle = _load_user_tokens().get(open_id)
    if isinstance(bundle, dict):
        name = str(bundle.get("name") or "").strip()
        if name:
            return name
    return _resolve_name_by_open_id(open_id) or _name_by_open_id(open_id) or open_id


def _format_manage_timestamp(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "-"
    return text.replace("T", " ")[:19]


def _default_auto_reply_settings() -> dict[str, Any]:
    return {
        "enabled": True,
        "busy_mode": False,
        "work_start": "09:00",
        "work_end": "18:30",
        "weekend_auto_reply": True,
        "cooldown_seconds": 600,
    }


def _default_auto_reply_faq() -> list[dict[str, Any]]:
    return [
        {
            "question": "上课时间",
            "patterns": ["上课时间", "什么时候上课", "几点上课", "课程时间", "直播时间", "开课时间"],
            "reply": "已收到。上课时间和课程安排需要按你当前报考项目确认。教务老师上线后会优先给你准确时间；如果你愿意，也可以直接补充“报考项目+年级/批次”，我们会更快核对。",
        },
        {
            "question": "报名材料",
            "patterns": ["报名材料", "需要什么材料", "报名资料", "要准备什么", "提交什么"],
            "reply": "已收到。常见报名材料通常包括身份证、学历证明、照片等，但不同项目要求会有差异。教务老师上线后会按你的项目给你发准确清单。",
        },
        {
            "question": "费用",
            "patterns": ["费用", "多少钱", "学费", "收费", "价格", "付款"],
            "reply": "已收到。费用和缴费方式需要按项目与服务内容确认。教务老师上线后会给你发对应项目的正式说明，避免口径不一致。",
        },
        {
            "question": "考试时间",
            "patterns": ["考试时间", "什么时候考试", "考试安排", "考试在什么时候", "考试日期"],
            "reply": "已收到。考试安排会按项目和批次确认。教务老师上线后会给你核对最近一次考试时间和准备节点。",
        },
        {
            "question": "证书毕业",
            "patterns": ["拿证", "毕业", "证书", "多久毕业", "多久拿证", "毕业时间"],
            "reply": "已收到。毕业/拿证时间和你的报考项目、进度有关，不能直接统一答复。教务老师上线后会按你的实际情况给你回复。",
        },
        {
            "question": "转人工",
            "patterns": ["人工", "老师在吗", "有人吗", "尽快回复", "联系老师", "需要老师回复"],
            "reply": "已收到，你的消息已经记录。当前教务老师可能暂时离线或消息较多，看到后会尽快人工回复你。",
        },
    ]


def _load_auto_reply_settings() -> dict[str, Any]:
    rows = _read_json(AUTO_REPLY_SETTING_FILE, {})
    if not isinstance(rows, dict):
        rows = {}
    merged = _default_auto_reply_settings()
    merged.update(rows)
    return merged


def _save_auto_reply_settings(rows: dict[str, Any]) -> None:
    merged = _default_auto_reply_settings()
    if isinstance(rows, dict):
        merged.update(rows)
    _write_json(AUTO_REPLY_SETTING_FILE, merged)


def _load_auto_reply_faq() -> list[dict[str, Any]]:
    rows = _read_json(AUTO_REPLY_FAQ_FILE, [])
    if not isinstance(rows, list) or not rows:
        rows = _default_auto_reply_faq()
        _write_json(AUTO_REPLY_FAQ_FILE, rows)
    return [row for row in rows if isinstance(row, dict)]


def _parse_hhmm(value: str, fallback_hour: int, fallback_minute: int) -> tuple[int, int]:
    try:
        hour_text, minute_text = (value or "").split(":", 1)
        hour = max(0, min(23, int(hour_text)))
        minute = max(0, min(59, int(minute_text)))
        return hour, minute
    except Exception:
        return fallback_hour, fallback_minute


def _should_auto_reply_now() -> tuple[bool, str]:
    rows = _load_auto_reply_settings()
    if not rows.get("enabled", True):
        return False, "disabled"
    if rows.get("busy_mode", False):
        return True, "busy_mode"
    now = datetime.now()
    if now.weekday() >= 5:
        return bool(rows.get("weekend_auto_reply", True)), "weekend"
    start_hour, start_minute = _parse_hhmm(str(rows.get("work_start", "09:00")), 9, 0)
    end_hour, end_minute = _parse_hhmm(str(rows.get("work_end", "18:30")), 18, 30)
    now_minutes = now.hour * 60 + now.minute
    start_minutes = start_hour * 60 + start_minute
    end_minutes = end_hour * 60 + end_minute
    if now_minutes < start_minutes or now_minutes > end_minutes:
        return True, "off_hours"
    return False, "work_hours"


def _can_send_auto_reply(sender_open_id: str) -> bool:
    settings = _load_auto_reply_settings()
    cooldown_seconds = int(settings.get("cooldown_seconds", 600) or 600)
    now_ts = time.time()
    rows = _read_json(AUTO_REPLY_STATE_FILE, {})
    if not isinstance(rows, dict):
        rows = {}
    last_ts = float(rows.get(sender_open_id, 0) or 0)
    if last_ts and now_ts - last_ts < cooldown_seconds:
        return False
    rows[sender_open_id] = now_ts
    _write_json(AUTO_REPLY_STATE_FILE, rows)
    return True


def _match_auto_reply_faq(text: str) -> dict[str, Any] | None:
    value = _norm_text(text)
    if not value:
        return None
    for row in _load_auto_reply_faq():
        patterns = row.get("patterns") or []
        if not isinstance(patterns, list):
            continue
        if any(str(p).strip() and str(p).strip() in value for p in patterns):
            return row
    return None


def _call_ai_auto_reply(student_name: str, text: str) -> str:
    if not (SETTINGS.ai_base_url and SETTINGS.ai_api_key and SETTINGS.ai_model):
        return ""
    faq_rows = _load_auto_reply_faq()
    faq_lines = []
    for row in faq_rows[:8]:
        question = str(row.get("question", "")).strip()
        reply = str(row.get("reply", "")).strip()
        if question and reply:
            faq_lines.append(f"- {question}: {reply}")
    system_prompt = (
        "你是招生教务机器人的离线值守助手。"
        "你的任务是在教务老师下班、周末或消息过多时，先自动回复学生的简单常见问题。"
        "只回答简单、低风险、低承诺的问题。"
        "不要承诺具体结果、不要伪造时间、不要给出无法确认的收费或考试结论。"
        "语气要简短、礼貌、稳定，结尾补一句“教务老师上线后会继续跟进”。"
        "如果问题涉及投诉、退费、举报、法律、情绪危机、高风险内容，只返回：TRANSFER_TO_HUMAN。"
    )
    user_prompt = (
        f"学生姓名：{student_name or '未登记'}\n"
        f"学生消息：{text}\n\n"
        f"可参考常见问答：\n{chr(10).join(faq_lines)}"
    )
    url = SETTINGS.ai_base_url.rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {SETTINGS.ai_api_key}",
        "Content-Type": "application/json",
    }
    body = {
        "model": SETTINGS.ai_model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.3,
        "max_tokens": 220,
    }
    try:
        resp = requests.post(url, headers=headers, json=body, timeout=20)
        data = resp.json()
        content = (((data.get("choices") or [{}])[0].get("message") or {}).get("content") or "").strip()
        if content == "TRANSFER_TO_HUMAN":
            return ""
        return content
    except Exception:
        return ""


def _build_auto_reply(student_name: str, text: str) -> tuple[str, str]:
    faq_hit = _match_auto_reply_faq(text)
    if faq_hit:
        return str(faq_hit.get("reply", "")).strip(), f"faq:{faq_hit.get('question', '')}"
    ai_reply = _call_ai_auto_reply(student_name, text)
    if ai_reply:
        return ai_reply, "ai"
    return "已收到你的消息。当前教务老师可能暂时离线，或正在集中处理消息。你的问题已经记录，教务老师上线后会继续跟进。", "fallback"


def _load_roster() -> dict[str, str]:
    return _read_json(ROSTER_FILE, {})


def _save_roster(roster: dict[str, str]) -> None:
    _write_json(ROSTER_FILE, roster)


def _load_authorized_users() -> dict[str, str]:
    raw = _read_json(AUTHORIZED_USER_FILE, {})
    rows: dict[str, str] = {}
    # 兼容历史格式：
    # 1) {"ou_xxx": "张三"}
    # 2) [{"open_id": "ou_xxx", "name": "张三"}]
    if isinstance(raw, dict):
        for open_id, name in raw.items():
            oid = str(open_id).strip()
            if oid:
                rows[oid] = str(name).strip()
    elif isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            oid = str(item.get("open_id") or "").strip()
            if not oid:
                continue
            rows[oid] = str(item.get("name") or "").strip()

    # 主管理员永远保留权限（不可丢失）
    teacher_id = (SETTINGS.teacher_open_id or "").strip()
    if _is_valid_open_id(teacher_id):
        rows.setdefault(teacher_id, _resolve_name_by_open_id(teacher_id) or "主管理员")
    return rows


def _save_authorized_users(rows: dict[str, str]) -> None:
    normalized: dict[str, str] = {}
    for open_id, name in rows.items():
        oid = str(open_id).strip()
        if oid:
            normalized[oid] = str(name).strip()
    # 强制保留主管理员权限，避免误删后失去管理能力
    teacher_id = (SETTINGS.teacher_open_id or "").strip()
    if _is_valid_open_id(teacher_id):
        normalized.setdefault(teacher_id, _resolve_name_by_open_id(teacher_id) or "主管理员")
    _write_json(AUTHORIZED_USER_FILE, normalized)


def _load_alert_receivers() -> dict[str, str]:
    raw = _read_json(ALERT_RECEIVER_FILE, {})
    if not isinstance(raw, dict):
        return {}
    rows: dict[str, str] = {}
    for open_id, name in raw.items():
        oid = str(open_id).strip()
        if oid:
            rows[oid] = str(name).strip()
    return rows


def _save_alert_receivers(rows: dict[str, str]) -> None:
    normalized: dict[str, str] = {}
    for open_id, name in rows.items():
        oid = str(open_id).strip()
        if oid:
            normalized[oid] = str(name).strip()
    _write_json(ALERT_RECEIVER_FILE, normalized)


def _resolve_name_by_open_id(open_id: str) -> str:
    org_rows = _read_json(ORG_CONTACT_FILE, [])
    for row in org_rows:
        if isinstance(row, dict) and (row.get("open_id") or "").strip() == open_id:
            return (row.get("name") or "").strip()
    # 兜底：再从本地 roster 查一次，避免与 _name_by_open_id 相互调用形成递归。
    roster = _load_roster()
    for name, oid in roster.items():
        if (oid or "").strip() == (open_id or "").strip() and (name or "").strip():
            return (name or "").strip()
    return ""


def _is_super_admin(open_id: str) -> bool:
    teacher_id = (SETTINGS.teacher_open_id or "").strip()
    return bool(open_id and teacher_id and open_id == teacher_id)


def _is_authorized_user(open_id: str) -> bool:
    if _is_super_admin(open_id):
        return True
    return open_id in _load_authorized_users()


def _grant_authorized_user(open_id: str, display_name: str = "") -> None:
    rows = _load_authorized_users()
    rows[open_id] = display_name.strip() or _resolve_name_by_open_id(open_id) or open_id
    _save_authorized_users(rows)


def _revoke_authorized_user(open_id: str) -> bool:
    # 主管理员权限不可移除
    if _is_super_admin(open_id):
        return False
    rows = _load_authorized_users()
    if open_id not in rows:
        return False
    rows.pop(open_id, None)
    _save_authorized_users(rows)
    return True


def _grant_alert_receiver(open_id: str, display_name: str = "") -> None:
    rows = _load_alert_receivers()
    rows[open_id] = display_name.strip() or _resolve_name_by_open_id(open_id) or open_id
    _save_alert_receivers(rows)


def _revoke_alert_receiver(open_id: str) -> bool:
    rows = _load_alert_receivers()
    if open_id not in rows:
        return False
    rows.pop(open_id, None)
    _save_alert_receivers(rows)
    return True


def _get_alert_receiver_open_ids() -> list[str]:
    rows = _load_alert_receivers()
    ids: list[str] = []
    teacher_id = (SETTINGS.teacher_open_id or "").strip()
    if _is_valid_open_id(teacher_id):
        ids.append(teacher_id)
    for open_id in rows.keys():
        if _is_valid_open_id(open_id) and open_id not in ids:
            ids.append(open_id)
    return ids


def _find_open_id_by_name(name: str) -> tuple[str, str]:
    target = (name or "").strip()
    if not target:
        return "", ""

    roster = _load_roster()
    exact = [(n, oid) for n, oid in roster.items() if (n or "").strip() == target]
    if len(exact) == 1:
        display_name, open_id = exact[0]
        return open_id, display_name

    compact_target = target.replace(" ", "")
    compact = [(n, oid) for n, oid in roster.items() if (n or "").replace(" ", "") == compact_target]
    if len(compact) == 1:
        display_name, open_id = compact[0]
        return open_id, display_name

    contains = [(n, oid) for n, oid in roster.items() if target in (n or "").strip() or (n or "").strip() in target]
    if len(contains) == 1:
        display_name, open_id = contains[0]
        return open_id, display_name

    org_rows = _read_json(ORG_CONTACT_FILE, [])
    org_matches = []
    for row in org_rows:
        if not isinstance(row, dict):
            continue
        row_name = (row.get("name") or "").strip()
        row_open_id = (row.get("open_id") or "").strip()
        if row_name and row_open_id and (row_name == target or target in row_name or row_name in target):
            org_matches.append((row_name, row_open_id))
    if len(org_matches) == 1:
        display_name, open_id = org_matches[0]
        return open_id, display_name

    return "", ""


def _find_member_candidates_by_name(name: str, limit: int = 10) -> list[dict[str, str]]:
    target = (name or "").strip()
    if not target:
        return []

    compact_target = target.replace(" ", "")
    seen_open_ids: set[str] = set()
    candidates: list[dict[str, str]] = []

    def add_candidate(row_name: str, row_open_id: str, source: str, priority: int) -> None:
        row_name = (row_name or "").strip()
        row_open_id = (row_open_id or "").strip()
        if not row_name or not row_open_id:
            return
        if row_open_id in seen_open_ids:
            return
        seen_open_ids.add(row_open_id)
        candidates.append(
            {
                "name": row_name,
                "open_id": row_open_id,
                "source": source,
                "priority": str(priority),
            }
        )

    def match_priority(row_name: str) -> int | None:
        row_name = (row_name or "").strip()
        if not row_name:
            return None
        if row_name == target:
            return 0
        if row_name.replace(" ", "") == compact_target:
            return 1
        if target in row_name or row_name in target:
            return 2
        return None

    roster = _load_roster()
    for row_name, row_open_id in roster.items():
        priority = match_priority(row_name)
        if priority is not None:
            add_candidate(row_name, row_open_id, "本地", priority)

    org_rows = _read_json(ORG_CONTACT_FILE, [])
    for row in org_rows:
        if not isinstance(row, dict):
            continue
        row_name = (row.get("name") or "").strip()
        row_open_id = (row.get("open_id") or "").strip()
        priority = match_priority(row_name)
        if priority is not None:
            add_candidate(row_name, row_open_id, "通讯录", priority)

    candidates.sort(key=lambda item: (int(item.get("priority", "99")), item["name"], item["open_id"]))
    return candidates[:limit]


def _hydrate_roster_by_names(
    names: list[str],
    *,
    auto_sync_if_empty: bool = False,
) -> tuple[list[str], list[str], int, str]:
    target_names = _dedup_keep_order([str(name).strip() for name in names if str(name).strip()])
    if not target_names:
        return [], [], 0, ""

    sync_added = 0
    sync_error = ""
    org_rows = _read_json(ORG_CONTACT_FILE, [])
    if auto_sync_if_empty and not org_rows and not SETTINGS.mock_send:
        sync_added, _, sync_error = _sync_org_contacts()

    roster = _load_roster()
    changed = False
    matched: list[str] = []
    missing: list[str] = []

    for name in target_names:
        if roster.get(name):
            matched.append(name)
            continue
        open_id, _ = _find_open_id_by_name(name)
        if open_id:
            roster[name] = open_id
            changed = True
            matched.append(name)
        else:
            missing.append(name)

    if changed:
        _save_roster(roster)
    return matched, missing, sync_added, sync_error


def _resolve_authorize_target(body: str, mentions: list[dict[str, Any]]) -> tuple[str, str]:
    value = (body or "").strip()
    if not value:
        return "", ""
    parts = value.split()
    if len(parts) == 1:
        mention_open_id = _extract_open_id_from_mentions(mentions)
        if mention_open_id:
            return mention_open_id, parts[0].strip()
        only = parts[0].strip()
        if only.startswith("ou_"):
            return only, _resolve_name_by_open_id(only)
        return _find_open_id_by_name(only)
    display_name = parts[0].strip()
    open_id = parts[1].strip()
    if open_id.startswith("ou_"):
        return open_id, display_name
    return "", ""


def _is_registration_command(text: str) -> bool:
    return text.startswith("学员注册 ") or text.startswith("瀛﹀憳娉ㄥ唽 ")


def _is_help_command(text: str) -> bool:
    return text in {"帮助", "幫助", "help", "菜单", "命令", "甯姪", "骞姪", "鑿滃崟", "鍛戒护"}


def _is_super_admin_command(text: str) -> bool:
    prefixes = [
        "授权使用 ",
        "移除使用权限 ",
        "授权接收预警 ",
        "移除预警接收 ",
    ]
    return any(text.startswith(p) for p in prefixes)


def _is_manager_command(text: str) -> bool:
    if _is_registration_command(text) or _is_help_command(text):
        return False
    if text in {"查看可用人员", "查看授权人员", "查看使用权限", "查看预警接收人", "查看预警权限"}:
        return True
    if text in {"开启自动回复", "关闭自动回复", "自动回复状态", "开启忙碌模式", "关闭忙碌模式"}:
        return True
    if text.startswith("设置自动回复时间 "):
        return True
    prefixes = [
        "同步组织成员",
        "同步成员",
        "同步通讯录",
        "搜索成员 ",
        "查看绑定",
        "查看群标签",
        "查看名单 ",
        "查看标签 ",
        "查看所有标签",
        "查看标签列表",
        "标签列表",
        "文件建标签 ",
        "添加到标签 ",
        "从标签移除 ",
        "重命名标签 ",
        "删除标签 ",
        "按标签群发 ",
        "设置预警阈值 ",
        "绑定学员 ",
        "保存名单 ",
        "绑定群 ",
        "记住本群 ",
        "复用群发 ",
        "群发 ",
        "群标签群发 ",
        "群成员私发 ",
        "按群私发 ",
        "保存群发记录 ",
        "按记录群发 ",
        "学生画像 ",
        "新建群发 ",
        "再发群发 ",
    ]
    return any(text.startswith(p) for p in prefixes)

def _load_student_tags() -> dict[str, list[str]]:
    raw = _read_json(STUDENT_TAG_FILE, {})
    if not isinstance(raw, dict):
        return {}
    tags: dict[str, list[str]] = {}
    for tag_name, names in raw.items():
        if not isinstance(tag_name, str):
            continue
        if isinstance(names, list):
            tags[tag_name] = [str(x).strip() for x in names if str(x).strip()]
    return tags


def _save_student_tags(tags: dict[str, list[str]]) -> None:
    normalized: dict[str, list[str]] = {}
    for tag_name, names in tags.items():
        key = str(tag_name).strip()
        if not key:
            continue
        normalized[key] = _dedup_keep_order([str(x).strip() for x in names if str(x).strip()])
    _write_json(STUDENT_TAG_FILE, normalized)


def _split_tag_names(value: str) -> list[str]:
    if not value:
        return []
    raw = value.replace("，", ",").replace("、", ",").replace("/", ",").replace("|", ",")
    return _dedup_keep_order([x.strip() for x in raw.split(",") if x.strip()])


def _bind_tags_for_student(student_name: str, tag_names: list[str]) -> list[str]:
    if not student_name:
        return []
    tags = _load_student_tags()
    added: list[str] = []
    for tag_name in _split_tag_names(",".join(tag_names)):
        names = tags.get(tag_name, [])
        if student_name not in names:
            names.append(student_name)
            tags[tag_name] = _dedup_keep_order(names)
        added.append(tag_name)
    _save_student_tags(tags)
    return _dedup_keep_order(added)


def _get_students_by_tag(tag_name: str) -> list[str]:
    tags = _load_student_tags()
    return _dedup_keep_order(tags.get(tag_name, []))


def _list_all_tag_names() -> list[str]:
    return sorted([x for x in _load_student_tags().keys() if x])


def _add_students_to_tag(tag_name: str, student_names: list[str]) -> list[str]:
    key = (tag_name or "").strip()
    if not key:
        return []
    tags = _load_student_tags()
    names = tags.get(key, [])
    names.extend([x.strip() for x in student_names if x and x.strip()])
    tags[key] = _dedup_keep_order(names)
    _save_student_tags(tags)
    return tags[key]


def _remove_students_from_tag(tag_name: str, student_names: list[str]) -> tuple[list[str], bool]:
    key = (tag_name or "").strip()
    if not key:
        return [], False
    tags = _load_student_tags()
    if key not in tags:
        return [], False
    remove_set = {x.strip() for x in student_names if x and x.strip()}
    tags[key] = [x for x in tags.get(key, []) if x not in remove_set]
    if not tags[key]:
        tags.pop(key, None)
        _save_student_tags(tags)
        return [], True
    _save_student_tags(tags)
    return tags[key], True


def _rename_student_tag(old_name: str, new_name: str) -> tuple[bool, str]:
    old_key = (old_name or "").strip()
    new_key = (new_name or "").strip()
    if not old_key or not new_key:
        return False, "标签名不能为空"
    tags = _load_student_tags()
    if old_key not in tags:
        return False, f"标签[{old_key}]不存在"
    merged = _dedup_keep_order(tags.get(new_key, []) + tags.get(old_key, []))
    tags[new_key] = merged
    if old_key != new_key:
        tags.pop(old_key, None)
    _save_student_tags(tags)
    return True, ""


def _delete_student_tag(tag_name: str) -> bool:
    key = (tag_name or "").strip()
    tags = _load_student_tags()
    if key not in tags:
        return False
    tags.pop(key, None)
    _save_student_tags(tags)
    return True


def _save_template_by_tag(template_name: str) -> tuple[int, int, list[str], list[str]]:
    names = _get_students_by_tag(template_name)
    if not names:
        return 0, 0, [], []
    return _save_template_by_names(template_name, names)


def _name_by_open_id(open_id: str) -> str:
    roster = _load_roster()
    matches: list[str] = []
    for name, oid in roster.items():
        if oid == open_id:
            matches.append(name)
    if not matches:
        return ""
    teacher_name = _resolve_name_by_open_id((SETTINGS.teacher_open_id or "").strip())
    for name in reversed(matches):
        if teacher_name and name == teacher_name:
            continue
        return name
    return matches[-1]


def _student_name_for_sender(open_id: str) -> str:
    return _resolve_name_by_open_id(open_id) or _name_by_open_id(open_id)


def _extract_open_id_from_mentions(mentions: list[dict[str, Any]]) -> str:
    for m in mentions or []:
        mid = m.get("id", {}) if isinstance(m, dict) else {}
        open_id = mid.get("open_id", "")
        if open_id:
            return open_id
    return ""


def _extract_mentions(mentions: list[dict[str, Any]]) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for m in mentions or []:
        if not isinstance(m, dict):
            continue
        mid = m.get("id", {}) if isinstance(m.get("id", {}), dict) else {}
        open_id = mid.get("open_id", "")
        name = (m.get("name") or "").strip()
        if open_id:
            rows.append((name, open_id))
    return rows


def _load_chat_tags() -> dict[str, str]:
    return _read_json(CHAT_TAG_FILE, {})


def _save_chat_tags(tags: dict[str, str]) -> None:
    _write_json(CHAT_TAG_FILE, tags)


def _load_last_send_ctx() -> dict[str, Any]:
    return _read_json(LAST_SEND_CTX_FILE, {})


def _save_last_send_ctx(ctx: dict[str, Any]) -> None:
    _write_json(LAST_SEND_CTX_FILE, ctx)


def _save_sender_last_context(sender_open_id: str, recipients: list[str], content: str, source: str) -> None:
    if not sender_open_id:
        return
    ctx = _load_last_send_ctx()
    ctx[sender_open_id] = {
        "recipients": recipients,
        "content": content,
        "source": source,
        "ts": datetime.now().isoformat(),
    }
    _save_last_send_ctx(ctx)


def _load_send_records() -> dict[str, Any]:
    return _read_json(SEND_RECORD_FILE, {})


def _save_send_records(records: dict[str, Any]) -> None:
    _write_json(SEND_RECORD_FILE, records)


def _dedup_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for v in values:
        if not v or v in seen:
            continue
        seen.add(v)
        out.append(v)
    return out


def _list_chat_member_open_ids(chat_id: str) -> tuple[list[str], str]:
    if not chat_id:
        return [], "chat_id 为空"
    if SETTINGS.mock_send:
        return [], "当前为 mock 模式，无法读取群成员"

    token = _token()
    headers = {"Authorization": f"Bearer {token}"}
    page_token = ""
    members: list[str] = []

    while True:
        params = {
            "page_size": 50,
            "member_id_type": "open_id",
        }
        if page_token:
            params["page_token"] = page_token
        resp = requests.get(
            f"{FEISHU_BASE}/im/v1/chats/{chat_id}/members",
            headers=headers,
            params=params,
            timeout=12,
        )
        data = resp.json()
        if data.get("code") != 0:
            msg = data.get("msg", "unknown")
            msg_l = str(msg).lower()
            if "scope" in msg_l or "permission" in msg_l or "access denied" in msg_l:
                return [], (
                    "读取群成员失败：缺少“会话/群成员读取”权限。\n"
                    "你当前开的“获取单聊、群组消息（im:message:readonly）”不等于“读取群信息/群成员”。\n"
                    "请在飞书开放平台权限管理里继续搜索并开通与以下 API 对应的读取权限：\n"
                    "1) im/v1/chats\n"
                    "2) im/v1/chats/{chat_id}/members"
                )
            return [], f"读取群成员失败：{msg}"

        block = data.get("data") or {}
        for item in block.get("items", []):
            member_id = (item.get("member_id") or "").strip()
            if member_id.startswith("ou_"):
                members.append(member_id)
        if not block.get("has_more"):
            break
        page_token = block.get("page_token", "")
        if not page_token:
            break

    members = _dedup_keep_order(members)
    if not members:
        return [], "群里没有可用成员（open_id）"
    return members, ""


def _find_chat_by_name(keyword: str) -> tuple[str, str, str]:
    if not keyword:
        return "", "", "群名称关键词为空"
    if SETTINGS.mock_send:
        return "", "", "当前为 mock 模式，无法按群名检索"

    token = _token()
    headers = {"Authorization": f"Bearer {token}"}
    page_token = ""
    matches: list[tuple[str, str]] = []
    keyword_l = keyword.lower()

    while True:
        # Keep params minimal for better compatibility across tenants.
        params = {
            "page_size": 50,
        }
        if page_token:
            params["page_token"] = page_token
        resp = requests.get(
            f"{FEISHU_BASE}/im/v1/chats",
            headers=headers,
            params=params,
            timeout=12,
        )
        data = resp.json()
        if data.get("code") != 0:
            msg = data.get("msg", "unknown")
            msg_l = str(msg).lower()
            if "scope" in msg_l or "permission" in msg_l or "access denied" in msg_l:
                return "", "", (
                    "按群名检索失败：缺少“会话信息读取”权限。\n"
                    "你现在开的群消息读取权限不够，还需要与以下 API 对应的读取权限：\n"
                    "1) im/v1/chats\n"
                    "2) im/v1/chats/{chat_id}/members"
                )
            if "field validation failed" in msg_l:
                return "", "", (
                    "按群名检索失败：参数校验失败。"
                    "请尝试使用更短关键词，或先在目标群里发送“记住本群 A组”。"
                )
            return "", "", f"按群名检索失败：{msg}"

        block = data.get("data") or {}
        for item in block.get("items", []):
            chat_id = (item.get("chat_id") or "").strip()
            name = (item.get("name") or "").strip()
            if not chat_id or not name:
                continue
            if keyword_l in name.lower():
                matches.append((chat_id, name))

        if len(matches) >= 5:
            break
        if not block.get("has_more"):
            break
        page_token = block.get("page_token", "")
        if not page_token:
            break

    if not matches:
        return "", "", f"未检索到群名包含“{keyword}”的会话（请确认机器人已在该群中）"
    if len(matches) > 1:
        name_list = "、".join(name for _, name in matches[:5])
        return "", "", f"匹配到多个群：{name_list}。请换一个更精确的群名关键词。"
    return matches[0][0], matches[0][1], ""


def _friendly_sync_error(msg: str, code: int | None = None) -> str:
    raw = (msg or "unknown").strip()
    lower = raw.lower()
    if "no dept authority" in lower:
        return (
            "同步失败：没有部门数据权限（no dept authority）。\n"
            "请到飞书开放平台 -> 应用 -> 权限管理：\n"
            "1) 开通 contact:user.base:readonly、contact:department.base:readonly\n"
            "2) 在“可访问的数据范围”里配置目标部门或全部成员\n"
            "3) 重新发布应用后再试“同步组织成员”"
        )
    if "permission" in lower or "access denied" in lower:
        return (
            f"同步失败：权限不足（{raw}）。\n"
            "请检查应用权限与数据范围配置，并发布后重试。"
        )
    if "tenant_access_token" in lower or (code is not None and code == 99991663):
        return "同步失败：应用凭证无效，请检查 FEISHU_APP_ID / FEISHU_APP_SECRET。"
    return f"同步失败：{raw}"


def _sync_org_contacts() -> tuple[int, int, str]:
    """
    Sync org users from Feishu and merge into local roster.
    Returns: (added_count, total_count, error_message)
    """
    if SETTINGS.mock_send:
        return 0, len(_load_roster()), "当前为 mock 模式，无法同步组织成员"

    token = _token()
    headers = {"Authorization": f"Bearer {token}"}
    users: list[dict[str, Any]] = []

    # Try department-based sync first.
    page_token = ""
    dept_sync_ok = True
    dept_err_msg = ""
    dept_err_code: int | None = None
    while True:
        params = {
            "department_id": SETTINGS.org_department_id,
            "user_id_type": "open_id",
            "page_size": 50,
            "fetch_child": "true",
        }
        if page_token:
            params["page_token"] = page_token

        resp = requests.get(
            f"{FEISHU_BASE}/contact/v3/users/find_by_department",
            headers=headers,
            params=params,
            timeout=12,
        )
        data = resp.json()
        if data.get("code") != 0:
            dept_sync_ok = False
            dept_err_msg = data.get("msg", "unknown")
            dept_err_code = data.get("code")
            break

        block = data.get("data") or {}
        users.extend(block.get("items", []))
        if not block.get("has_more"):
            break
        page_token = block.get("page_token", "")
        if not page_token:
            break

    # Fallback: when dept authority is missing, use visible user list API.
    if not dept_sync_ok and "no dept authority" in dept_err_msg.lower():
        users = []
        page_token = ""
        while True:
            params = {
                "user_id_type": "open_id",
                "page_size": 50,
            }
            if page_token:
                params["page_token"] = page_token
            resp = requests.get(
                f"{FEISHU_BASE}/contact/v3/users",
                headers=headers,
                params=params,
                timeout=12,
            )
            data = resp.json()
            if data.get("code") != 0:
                return 0, len(_load_roster()), _friendly_sync_error(data.get("msg", "unknown"), data.get("code"))
            block = data.get("data") or {}
            users.extend(block.get("items", []))
            if not block.get("has_more"):
                break
            page_token = block.get("page_token", "")
            if not page_token:
                break
    elif not dept_sync_ok:
        return 0, len(_load_roster()), _friendly_sync_error(dept_err_msg, dept_err_code)

    roster = _load_roster()
    before = len(roster)
    rows: list[dict[str, str]] = []
    for u in users:
        name = (u.get("name") or "").strip()
        open_id = (u.get("open_id") or "").strip()
        if not name or not open_id:
            continue
        roster[name] = open_id
        rows.append({"name": name, "open_id": open_id})

    _save_roster(roster)
    _write_json(ORG_CONTACT_FILE, rows)
    return max(0, len(roster) - before), len(roster), ""


def _allow_alert(level: str) -> bool:
    return _risk_score(level) >= _risk_score(SETTINGS.alert_level)


def _analyze_student(name: str, text: str) -> dict[str, Any]:
    lower = text.lower()
    risk = "低"
    sentiment = "中性"
    tags: list[str] = []
    advice = "保持日常跟进，确认学习节奏和目标稳定。"

    high_keywords = ["焦虑", "压力", "崩溃", "不想学", "退学", "失眠", "抑郁", "绝望", "扛不住", "想放弃"]
    mid_keywords = ["犹豫", "纠结", "不确定", "担心", "迷茫", "拖延", "疲惫", "没状态"]
    pos_keywords = ["积极", "准备", "目标", "想报名", "愿意", "推进", "执行"]

    complaint_mid_keywords = [
        "不满", "反感", "生气", "离谱", "没说清楚", "说清楚", "不负责", "负责不负责",
        "官方", "跟机器人一样", "没完了", "不要再发", "别再发", "重复发", "营销短信",
        "不需要其他服务", "不想搞其他事情", "不找你们了", "别联系我", "停止服务",
        "根本不行", "不合格", "差得离谱", "完全没用", "我受够了", "没法接受",
        "做的真差", "太失望了", "真让人生气", "看不下去", "一点都不靠谱", "白费力气",
        "无能", "你们不靠谱", "不想再接触你们", "我不满意", "态度差", "服务差劲",
        "没效果", "太无知了", "去找别人", "你们完全不理解我", "不想听了", "没有兴趣了",
        "已经没有耐心了", "没有动力了", "烦死了", "很失望", "有点心烦", "太累了", "没意思了",
        "不想继续聊下去了", "不想再听你们推销了", "不要再骚扰我", "不感兴趣", "浪费我的时间",
        "不可信", "让我很担心", "没保障", "不敢再继续了", "不想再参与了", "不要再打扰我",
        "不想接受你们的帮助", "不想继续了", "不想听你说", "我不在乎", "不想做", "不想再了解",
        "你们太慢了", "怎么这么慢", "什么都没有解决", "为什么拖这么久", "没有解决我的问题",
        "不值得信赖", "实际做不到", "根本不专业", "得不到帮助", "都无所谓了", "没什么好说的",
        "懒得理你们", "随便吧", "家里人不让我做这个", "家庭压力太大了", "家里没钱", "家里有很多困难",
    ]
    complaint_high_keywords = [
        "坑", "欺骗", "诱导消费", "欺骗诱导消费", "找你们", "什么渠道", "投诉", "维权",
        "拿不到证书", "会找你们", "太离谱了", "没有你们这么做的",
    ]
    exposure_keywords = [
        "举报", "曝光", "差评", "报警", "消费者协会", "找律师", "起诉", "法院", "法律途径",
        "法律援助", "媒体曝光", "微博", "抖音", "小红书", "公开这个问题", "行业协会", "大v",
        "社交平台", "网上投诉", "相关部门投诉", "向领导举报",
    ]
    churn_keywords = [
        "找其他人", "换地方", "不会再联系你们", "再也不找你们了", "不再和你们合作了",
        "停止和你们合作", "不再和你们沟通了", "换人帮我", "去别的平台", "去找别人处理",
        "我不会再联系你们了", "我决定停止和你们合作", "去别的地方试试看", "不打算再和你们合作了",
    ]
    distrust_keywords = [
        "不相信你们", "感觉被欺骗了", "太可怕了", "不敢相信", "担心被骗", "不可信",
        "真不可信", "怎么这么不靠谱", "不安", "是不是在瞒我", "怀疑你们有问题",
    ]
    family_pressure_keywords = [
        "家庭不支持", "父母不同意", "家里总是反对我", "家里情况不允许", "家里没钱",
        "被家里逼迫了", "家里有很多困难", "家人不同意我的选择",
    ]
    urgency_keywords = ["马上", "立刻", "尽快", "今天就", "现在就"]

    hit_signals: list[str] = []

    high_hits = [k for k in high_keywords if k in lower]
    complaint_high_hits = [k for k in complaint_high_keywords if k in lower]
    complaint_mid_hits = [k for k in complaint_mid_keywords if k in lower]
    exposure_hits = [k for k in exposure_keywords if k in lower]
    churn_hits = [k for k in churn_keywords if k in lower]
    distrust_hits = [k for k in distrust_keywords if k in lower]
    family_pressure_hits = [k for k in family_pressure_keywords if k in lower]
    mid_hits = [k for k in mid_keywords if k in lower]
    pos_hits = [k for k in pos_keywords if k in lower]
    urgency_hits = [k for k in urgency_keywords if k in lower]

    if high_hits:
        risk = "高"
        sentiment = "负向"
        tags = ["高压", "情绪波动", "需要陪伴"]
        advice = "建议今天进行一对一沟通，先安抚情绪，再给出可执行的小目标。"
        hit_signals.extend(high_hits[:3])
    elif complaint_high_hits:
        risk = "高"
        sentiment = "强负向"
        tags = ["强烈不满", "投诉风险", "流失风险", "需要安抚"]
        advice = "建议立即人工接手沟通，先确认核心诉求与责任边界，停止重复营销触达，并在当天给出明确处理时间。"
        hit_signals.extend(complaint_high_hits[:3])
    elif exposure_hits:
        risk = "高"
        sentiment = "强负向"
        tags = ["举报曝光风险", "舆情风险", "需要升级处理"]
        advice = "建议立即升级主管处理，统一对外口径，暂停争辩式回复，并保留完整沟通记录。"
        hit_signals.extend(exposure_hits[:3])
    elif churn_hits:
        risk = "高"
        sentiment = "强负向"
        tags = ["流失风险", "合作终止倾向", "需要挽回"]
        advice = "建议先停止施压式触达，确认对方是否仍愿意沟通，再由人工给出单点解决方案。"
        hit_signals.extend(churn_hits[:3])
    elif complaint_mid_hits:
        risk = "中"
        sentiment = "偏负向"
        tags = ["明显反感", "服务摩擦", "需要安抚"]
        advice = "建议减少模板化回复，先共情确认不满点，再给出一条明确处理方案和下次反馈时间。"
        hit_signals.extend(complaint_mid_hits[:3])
    elif distrust_hits:
        risk = "中"
        sentiment = "偏负向"
        tags = ["不信任", "需要解释澄清"]
        advice = "建议减少口头承诺，改为明确流程、凭据和时间节点，优先恢复信任感。"
        hit_signals.extend(distrust_hits[:3])
    elif family_pressure_hits:
        risk = "中"
        sentiment = "偏负向"
        tags = ["外部压力", "家庭阻力", "需要安抚"]
        advice = "建议先了解家庭或经济压力来源，避免直接催促推进，给出更低压力的跟进路径。"
        hit_signals.extend(family_pressure_hits[:3])
    elif mid_hits:
        risk = "中"
        sentiment = "偏负向"
        tags = ["决策迟疑", "需要引导"]
        advice = "建议提供2-3个清晰方案，用时间节点推动决策。"
        hit_signals.extend(mid_hits[:3])
    elif pos_hits:
        risk = "低"
        sentiment = "正向"
        tags = ["目标明确", "执行意愿高"]
        advice = "建议推进报名关键动作，设置本周完成节点。"
        hit_signals.extend(pos_hits[:3])

    if not tags:
        tags = ["常规跟进"]

    if urgency_hits and risk != "低":
        tags = list(dict.fromkeys(tags + ["需要尽快处理"]))
        hit_signals.extend(urgency_hits[:2])

    if any(k in lower for k in ["不要再发", "别再发", "营销短信", "没完了", "重复发"]):
        tags = list(dict.fromkeys(tags + ["营销触达敏感"]))
        if risk == "低":
            risk = "中"
            sentiment = "偏负向"
            advice = "建议暂停重复催促类消息，改为一次性说明后等待学生主动回复。"

    if any(k in lower for k in ["拿不到证书", "找你们", "投诉", "维权", "什么渠道", "举报", "曝光", "起诉", "报警"]):
        tags = list(dict.fromkeys(tags + ["升级处理风险"]))
        risk = "高"
        sentiment = "强负向"
        advice = "建议立即升级主管介入，统一口径说明处理路径，并保留后续跟进记录。"

    all_profiles = _read_json(PROFILE_FILE, [])
    history = [p for p in all_profiles if p.get("student_name") == name]
    high_risk_count = sum(1 for p in history if p.get("risk_level") == "高")

    if risk == "高" and high_risk_count >= 2:
        advice = "连续高风险预警，建议立即升级为重点跟进：当天1v1沟通+家校同步+48小时复盘。"
    elif risk == "中" and high_risk_count >= 1:
        advice = "近期波动上升，建议缩短跟进周期到每天一次，并给出明确时间节点。"

    payload = {
        "student_name": name,
        "sentiment": sentiment,
        "risk_level": risk,
        "personality_tags": tags,
        "advice": advice,
        "source_text": text,
        "history_count": len(history),
        "high_risk_count": high_risk_count,
        "hit_signals": list(dict.fromkeys(hit_signals))[:5],
    }
    all_profiles.append(payload)
    _write_json(PROFILE_FILE, all_profiles)
    return payload


def _build_profile_card(profile: dict[str, Any]) -> dict[str, Any]:
    return {
        "config": {"wide_screen_mode": True},
        "header": {"template": "red" if profile["risk_level"] == "高" else "blue", "title": {"tag": "plain_text", "content": f"学生预警 | {profile['student_name']}"}},
        "elements": [
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**情感倾向**：{profile['sentiment']}"}},
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**风险等级**：{profile['risk_level']}"}},
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**性格标签**：{'、'.join(profile['personality_tags'])}"}},
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**触发信号**：{'、'.join(profile.get('hit_signals') or ['无'])}"}},
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**历史跟进次数**：{profile.get('history_count', 0)}"}},
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**建议反馈**：{profile['advice']}"}},
            {"tag": "hr"},
            {"tag": "note", "elements": [{"tag": "plain_text", "content": "仅老师可见：此卡片用于预警与跟进建议。"}]},
        ],
    }


def _save_template_by_names(template_name: str, names: list[str]) -> tuple[int, int, list[str], list[str]]:
    _hydrate_roster_by_names(names, auto_sync_if_empty=True)
    roster = _load_roster()
    ok_ids: list[str] = []
    missing: list[str] = []
    for n in names:
        open_id = roster.get(n)
        if open_id:
            ok_ids.append(open_id)
        else:
            missing.append(n)

    rows = _read_json(TEMPLATE_FILE, [])
    rows = [r for r in rows if r.get("template_name") != template_name]
    rows.append({"template_name": template_name, "recipient_open_ids": ok_ids, "recipient_names": names})
    _write_json(TEMPLATE_FILE, rows)
    return len(ok_ids), len(names), names, missing


def _send_with_tag(tag_name: str, content: str) -> dict[str, Any]:
    names = _get_students_by_tag(tag_name)
    if not names:
        return {"ok": False, "msg": f"标签[{tag_name}]下还没有学员"}
    ok_count, total_count, _, missing = _save_template_by_names(tag_name, names)
    result = _send_with_template(tag_name, content)
    result["tag_name"] = tag_name
    result["recipient_names"] = names
    result["matched"] = ok_count
    result["expected_total"] = total_count
    result["missing_names"] = missing
    return result


def _send_with_template(template_name: str, content: str) -> dict[str, Any]:
    rows = _read_json(TEMPLATE_FILE, [])
    row = next((r for r in rows if r.get("template_name") == template_name), None)
    if not row:
        return {"ok": False, "msg": "模板不存在"}
    recipients = row.get("recipient_open_ids", [])
    if not recipients:
        return {"ok": False, "msg": "模板中没有有效接收人"}

    logs = _read_json(SEND_LOG_FILE, [])
    success = 0
    items: list[dict[str, Any]] = []
    for open_id in recipients:
        ret = _send_text(open_id, content)
        ok = ret.get("code") == 0
        if ok:
            success += 1
        items.append({"open_id": open_id, "ok": ok, "raw": ret})
        logs.append({"template_name": template_name, "open_id": open_id, "content": content, "ok": ok, "raw": ret})
    _write_json(SEND_LOG_FILE, logs)
    return {"ok": True, "template_name": template_name, "total": len(recipients), "success": success, "results": items}


class TemplateUpsert(BaseModel):
    template_name: str = Field(..., description="模板名称")
    recipient_open_ids: list[str] = Field(..., description="学员 open_id 列表")


class SendRequest(BaseModel):
    template_name: str
    content: str


class EventEnvelope(BaseModel):
    type: str | None = None
    challenge: str | None = None
    token: str | None = None
    event: dict[str, Any] | None = None
    header: dict[str, Any] | None = None


app = FastAPI(title="Feishu Enrollment Bot MVP")


@app.on_event("startup")
def _startup() -> None:
    env_values = _load_dotenv()
    SETTINGS.app_id = env_values.get("FEISHU_APP_ID", "")
    SETTINGS.app_secret = env_values.get("FEISHU_APP_SECRET", "")
    SETTINGS.verify_token = env_values.get("FEISHU_VERIFY_TOKEN", "")
    SETTINGS.teacher_open_id = env_values.get("TEACHER_OPEN_ID", "")
    SETTINGS.org_department_id = env_values.get("ORG_DEPARTMENT_ID", "0") or "0"
    SETTINGS.alert_level = env_values.get("ALERT_LEVEL", "中") or "中"
    SETTINGS.mock_send = not (SETTINGS.app_id and SETTINGS.app_secret)
    SETTINGS.ai_base_url = env_values.get("AI_BASE_URL", "")
    SETTINGS.ai_api_key = env_values.get("AI_API_KEY", "")
    SETTINGS.ai_model = env_values.get("AI_MODEL", "")
    SETTINGS.public_base_url = env_values.get("PUBLIC_BASE_URL", "").strip().rstrip("/")
    SETTINGS.oauth_state_secret = env_values.get("OAUTH_STATE_SECRET", "")
    _load_auto_reply_faq()
    _save_auto_reply_settings(_load_auto_reply_settings())


@app.get("/healthz")
def healthz() -> dict[str, Any]:
    return {"ok": True, "mock_send": SETTINGS.mock_send, "alert_level": SETTINGS.alert_level}


@app.get("/auth/feishu/login")
def feishu_oauth_login() -> RedirectResponse:
    if not SETTINGS.public_base_url:
        raise HTTPException(status_code=400, detail="缺少 PUBLIC_BASE_URL，请先配置固定域名")
    if not SETTINGS.app_id or not SETTINGS.app_secret:
        raise HTTPException(status_code=400, detail="缺少 FEISHU_APP_ID / FEISHU_APP_SECRET")
    return RedirectResponse(url=_oauth_login_url(), status_code=302)


@app.get("/auth/feishu/callback")
def feishu_oauth_callback(code: str = Query(default=""), state: str = Query(default="")) -> Any:
    if not code:
        raise HTTPException(status_code=400, detail="回调缺少 code")
    state_payload = _decode_state_blob(state)
    if not state_payload:
        raise HTTPException(status_code=400, detail="OAuth state 校验失败")
    token_data = _exchange_user_access_token(code)
    user_info = _fetch_user_identity(token_data.get("access_token", ""))
    saved = _store_user_token_bundle(token_data, user_info)
    teacher_open_id = (SETTINGS.teacher_open_id or "").strip()
    purpose = str(state_payload.get("purpose") or "login")
    next_path = str(state_payload.get("next") or "/manage")
    if purpose == "manage":
        if teacher_open_id and teacher_open_id != saved.get("open_id"):
            raise HTTPException(status_code=403, detail="只有教师账号可以登录管理页")
        response = RedirectResponse(url=next_path or "/manage", status_code=302)
        response.set_cookie(
            key=MANAGE_SESSION_COOKIE,
            value=_build_manage_session(saved.get("open_id", "")),
            max_age=MANAGE_SESSION_MAX_AGE,
            httponly=True,
            secure=True,
            samesite="lax",
            path="/",
        )
        return response
    return {
        "ok": True,
        "message": "飞书用户授权成功",
        "authorized_open_id": saved.get("open_id"),
        "authorized_name": saved.get("name"),
        "is_teacher_match": bool(teacher_open_id and teacher_open_id == saved.get("open_id")),
        "next_step": "如果 is_teacher_match 为 false，请检查 .env 里的 TEACHER_OPEN_ID 是否正确。",
    }


@app.get("/auth/feishu/status")
def feishu_oauth_status() -> dict[str, Any]:
    bundle = _get_teacher_user_token_bundle()
    teacher_open_id = (SETTINGS.teacher_open_id or "").strip()
    return {
        "ok": True,
        "teacher_open_id": teacher_open_id,
        "authorized": bool(bundle),
        "authorized_name": (bundle or {}).get("name", ""),
        "redirect_uri": f"{SETTINGS.public_base_url}/auth/feishu/callback" if SETTINGS.public_base_url else "",
        "login_url": f"{SETTINGS.public_base_url}/auth/feishu/login" if SETTINGS.public_base_url else "",
    }


def _build_manage_state(open_id: str) -> dict[str, Any]:
    tags_raw = _load_student_tags()
    tags = [
        {"name": tag_name, "members": names, "count": len(names)}
        for tag_name, names in sorted(tags_raw.items(), key=lambda item: item[0])
    ]

    authorized_raw = _load_authorized_users()
    authorized_users = [
        {"open_id": oid, "name": name}
        for oid, name in sorted(authorized_raw.items(), key=lambda item: (item[1] or item[0]).lower())
    ]

    alert_raw = _load_alert_receivers()
    alert_receivers = [
        {"open_id": oid, "name": name}
        for oid, name in sorted(alert_raw.items(), key=lambda item: (item[1] or item[0]).lower())
    ]

    send_records_raw = _read_json(SEND_RECORD_FILE, {})
    send_records: list[dict[str, Any]] = []
    if isinstance(send_records_raw, dict):
        for record_name, row in send_records_raw.items():
            if not isinstance(row, dict):
                continue
            recipients = row.get("recipients", [])
            if not isinstance(recipients, list):
                recipients = []
            send_records.append(
                {
                    "record_name": str(record_name),
                    "owner": str(row.get("owner") or "").strip(),
                    "source": str(row.get("source") or "").strip(),
                    "saved_at": str(row.get("saved_at") or "").strip(),
                    "last_content": str(row.get("last_content") or "").strip(),
                    "recipient_count": len([x for x in recipients if str(x).strip()]),
                    "recipients": [str(x).strip() for x in recipients if str(x).strip()],
                }
            )
    send_records.sort(key=lambda item: item.get("saved_at", ""), reverse=True)

    send_logs_raw = _read_json(SEND_LOG_FILE, [])
    send_logs: list[dict[str, Any]] = []
    if isinstance(send_logs_raw, list):
        for row in send_logs_raw[-30:]:
            if isinstance(row, dict):
                send_logs.append(row)

    templates_raw = _read_json(TEMPLATE_FILE, [])
    templates: list[dict[str, Any]] = []
    if isinstance(templates_raw, list):
        for row in templates_raw:
            if not isinstance(row, dict):
                continue
            recipient_open_ids = row.get("recipient_open_ids", [])
            recipient_names = row.get("recipient_names", [])
            if not isinstance(recipient_open_ids, list):
                recipient_open_ids = []
            if not isinstance(recipient_names, list):
                recipient_names = []
            templates.append(
                {
                    "template_name": str(row.get("template_name") or "").strip(),
                    "recipient_count": len([x for x in recipient_open_ids if str(x).strip()]),
                    "recipient_names": [str(x).strip() for x in recipient_names if str(x).strip()],
                }
            )
    templates.sort(key=lambda item: item.get("template_name", ""))

    roster = _load_roster()
    org_contacts = _read_json(ORG_CONTACT_FILE, [])
    last_uploaded = _load_last_uploaded_name_rows()
    teacher_name = _display_name_for_open_id(open_id)
    return {
        "ok": True,
        "current_user": {
            "open_id": open_id,
            "name": teacher_name,
            "teacher_open_id": (SETTINGS.teacher_open_id or "").strip(),
        },
        "summary": {
            "tag_count": len(tags),
            "authorized_count": len(authorized_users),
            "alert_receiver_count": len(alert_receivers),
            "record_count": len(send_records),
            "template_count": len(templates),
            "contact_count": len(org_contacts) if isinstance(org_contacts, list) else 0,
            "roster_count": len(roster),
            "last_upload_count": len(last_uploaded),
        },
        "tags": tags,
        "authorized_users": authorized_users,
        "alert_receivers": alert_receivers,
        "send_records": send_records,
        "templates": templates,
        "recent_send_logs": send_logs,
    }


def _restart_service_async() -> None:
    command = (
        f"sleep 1; "
        f"(systemctl restart {MANAGE_SERVICE_NAME} >/tmp/{MANAGE_SERVICE_NAME}.restart.log 2>&1 "
        f"|| sudo -n systemctl restart {MANAGE_SERVICE_NAME} >/tmp/{MANAGE_SERVICE_NAME}.restart.log 2>&1)"
    )
    subprocess.Popen(
        ["bash", "-lc", command],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )


def _render_manage_page() -> str:
    return """
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>招生助手管理台</title>
  <style>
    :root {
      --bg: #f5f7fb;
      --card: rgba(255,255,255,.92);
      --line: rgba(15,23,42,.08);
      --text: #0f172a;
      --muted: #64748b;
      --primary: #2563eb;
      --primary2: #0ea5e9;
      --success: #16a34a;
      --warning: #d97706;
      --shadow: 0 18px 50px rgba(15,23,42,.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(37,99,235,.12), transparent 30%),
        radial-gradient(circle at right top, rgba(14,165,233,.12), transparent 24%),
        linear-gradient(180deg, #fbfdff 0%, #eef4ff 100%);
      min-height: 100vh;
    }
    .shell {
      max-width: 1440px;
      margin: 0 auto;
      padding: 28px 20px 48px;
    }
    .hero {
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: center;
      padding: 22px 24px;
      border: 1px solid var(--line);
      border-radius: 24px;
      background: linear-gradient(135deg, rgba(255,255,255,.95), rgba(245,249,255,.92));
      box-shadow: var(--shadow);
      backdrop-filter: blur(12px);
    }
    .title {
      margin: 0 0 6px 0;
      font-size: 30px;
      letter-spacing: .5px;
    }
    .subtitle {
      margin: 0;
      color: var(--muted);
      line-height: 1.7;
    }
    .hero-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      justify-content: flex-end;
    }
    .btn {
      border: 0;
      border-radius: 999px;
      padding: 12px 18px;
      font-size: 14px;
      font-weight: 600;
      cursor: pointer;
      transition: transform .15s ease, box-shadow .15s ease, opacity .15s ease;
    }
    .btn:hover { transform: translateY(-1px); }
    .btn-primary { color: #fff; background: linear-gradient(135deg, var(--primary), var(--primary2)); box-shadow: 0 10px 24px rgba(37,99,235,.18); }
    .btn-ghost { color: var(--text); background: #fff; border: 1px solid var(--line); }
    .notice {
      margin-top: 14px;
      padding: 14px 18px;
      border-radius: 16px;
      background: rgba(255,255,255,.82);
      border: 1px solid var(--line);
      color: var(--muted);
    }
    .grid {
      margin-top: 18px;
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 14px;
    }
    .stat {
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 18px;
      box-shadow: var(--shadow);
      min-height: 104px;
    }
    .stat-label {
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 10px;
    }
    .stat-value {
      font-size: 28px;
      font-weight: 800;
      margin-bottom: 8px;
    }
    .stat-desc {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.5;
    }
    .sections {
      margin-top: 18px;
      display: grid;
      grid-template-columns: 1.15fr .85fr;
      gap: 14px;
    }
    .panel {
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 22px;
      box-shadow: var(--shadow);
      padding: 18px;
      min-height: 220px;
    }
    .panel h2 {
      margin: 0 0 14px 0;
      font-size: 18px;
    }
    .panel-head {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      margin-bottom: 12px;
    }
    .panel-head .small {
      color: var(--muted);
      font-size: 13px;
    }
    .list {
      display: grid;
      gap: 10px;
    }
    .tag-card {
      border: 1px solid var(--line);
      border-radius: 16px;
      padding: 12px 14px;
      background: #fff;
    }
    .tag-card summary {
      cursor: pointer;
      font-weight: 700;
      list-style: none;
    }
    .tag-card summary::-webkit-details-marker { display: none; }
    .chip-wrap {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 10px;
    }
    .chip {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 6px 10px;
      border-radius: 999px;
      background: #eef4ff;
      color: #1d4ed8;
      font-size: 12px;
      line-height: 1;
    }
    .chip.gray { background: #f1f5f9; color: #334155; }
    .table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    .table th, .table td {
      border-bottom: 1px solid var(--line);
      padding: 10px 8px;
      vertical-align: top;
      text-align: left;
    }
    .table th { color: var(--muted); font-weight: 700; }
    .table td small { color: var(--muted); display: block; margin-top: 4px; }
    .two-col {
      margin-top: 14px;
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }
    .mono { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; }
    .status {
      margin-top: 16px;
      border-radius: 16px;
      padding: 12px 14px;
      background: #fff;
      border: 1px solid var(--line);
      color: var(--muted);
      min-height: 48px;
    }
    .ok { color: var(--success); font-weight: 700; }
    .err { color: #dc2626; font-weight: 700; }
    @media (max-width: 1200px) {
      .grid, .sections, .two-col { grid-template-columns: 1fr 1fr; }
    }
    @media (max-width: 860px) {
      .hero, .sections, .two-col, .grid { grid-template-columns: 1fr; display: grid; }
      .hero-actions { justify-content: flex-start; }
    }
  </style>
</head>
<body>
  <div class="shell">
    <div class="hero">
      <div>
        <h1 class="title">招生助手管理台</h1>
        <p class="subtitle">
          浏览标签、授权人员、发送记录，并支持一键同步组织成员和重启服务。<br>
          仅限管理员登录后访问。
        </p>
      </div>
      <div class="hero-actions">
        <button class="btn btn-primary" id="syncBtn">一键同步通讯录</button>
        <button class="btn btn-primary" id="restartBtn">一键重启服务</button>
        <button class="btn btn-ghost" id="refreshBtn">刷新页面</button>
        <button class="btn btn-ghost" id="logoutBtn">退出登录</button>
      </div>
    </div>

    <div class="notice" id="userInfo">正在加载登录信息...</div>

    <div class="grid" id="stats"></div>

    <div class="sections">
      <div class="panel">
        <div class="panel-head">
          <h2>标签</h2>
          <div class="small" id="tagHint">正在加载...</div>
        </div>
        <div class="list" id="tagList"></div>
      </div>
      <div class="panel">
        <div class="panel-head">
          <h2>授权人员</h2>
          <div class="small" id="authHint">正在加载...</div>
        </div>
        <div class="chip-wrap" id="authList"></div>
      </div>
    </div>

    <div class="two-col">
      <div class="panel">
        <div class="panel-head">
          <h2>发送记录</h2>
          <div class="small">最近保存的群发模板 / 记录</div>
        </div>
        <table class="table">
          <thead>
            <tr>
              <th>名称</th>
              <th>接收人</th>
              <th>最近内容</th>
              <th>时间</th>
            </tr>
          </thead>
          <tbody id="recordTable"></tbody>
        </table>
      </div>
      <div class="panel">
        <div class="panel-head">
          <h2>最近发送日志</h2>
          <div class="small" id="logHint">最近 30 条</div>
        </div>
        <table class="table">
          <thead>
            <tr>
              <th>时间</th>
              <th>模板</th>
              <th>结果</th>
            </tr>
          </thead>
          <tbody id="logTable"></tbody>
        </table>
      </div>
    </div>

    <div class="status" id="statusBox">准备就绪。</div>
  </div>

  <script>
    const stateUrl = "/manage/api/state";
    const syncUrl = "/manage/api/sync-org-contacts";
    const restartUrl = "/manage/api/restart-service";

    const statusBox = document.getElementById("statusBox");
    const userInfo = document.getElementById("userInfo");
    const statsEl = document.getElementById("stats");
    const tagList = document.getElementById("tagList");
    const tagHint = document.getElementById("tagHint");
    const authList = document.getElementById("authList");
    const authHint = document.getElementById("authHint");
    const recordTable = document.getElementById("recordTable");
    const logTable = document.getElementById("logTable");

    function esc(value) {
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#39;");
    }

    function setStatus(msg, kind = "ok") {
      statusBox.innerHTML = kind === "err" ? `<span class="err">${esc(msg)}</span>` : `<span class="ok">${esc(msg)}</span>`;
    }

    async function api(url, options = {}) {
      const res = await fetch(url, {
        credentials: "same-origin",
        headers: {"Content-Type": "application/json", ...(options.headers || {})},
        ...options,
      });
      if (res.status === 401) {
        window.location.href = "/manage/login";
        return null;
      }
      const data = await res.json().catch(() => ({}));
      if (!res.ok || data.ok === false) {
        const msg = data.detail || data.message || data.msg || `请求失败 (${res.status})`;
        throw new Error(msg);
      }
      return data;
    }

    function renderStats(summary) {
      const cards = [
        ["标签", summary.tag_count, "当前保存的学生标签数量"],
        ["授权人员", summary.authorized_count, "可使用群发/管理功能的人员"],
        ["发送记录", summary.record_count, "已保存的群发模板或记录"],
        ["通讯录", summary.contact_count, "同步到本地的组织成员数量"],
      ];
      statsEl.innerHTML = cards.map(([label, value, desc]) => `
        <div class="stat">
          <div class="stat-label">${esc(label)}</div>
          <div class="stat-value">${esc(value)}</div>
          <div class="stat-desc">${esc(desc)}</div>
        </div>
      `).join("");
    }

    function renderTags(tags) {
      tagHint.textContent = `${tags.length} 个标签`;
      if (!tags.length) {
        tagList.innerHTML = '<div class="chip gray">当前还没有标签。</div>';
        return;
      }
      tagList.innerHTML = tags.map(tag => {
        const members = (tag.members || []).map(name => `<span class="chip gray">${esc(name)}</span>`).join("");
        return `
          <details class="tag-card">
            <summary>${esc(tag.name)} <span class="chip">${esc(tag.count)}</span></summary>
            <div class="chip-wrap">${members || '<span class="chip gray">暂无成员</span>'}</div>
          </details>
        `;
      }).join("");
    }

    function renderAuth(users) {
      authHint.textContent = `${users.length} 人`;
      if (!users.length) {
        authList.innerHTML = '<span class="chip gray">当前没有额外授权人员。</span>';
        return;
      }
      authList.innerHTML = users.map(item => `
        <span class="chip">${esc(item.name || item.open_id)}</span>
      `).join("");
    }

    function renderRecords(records) {
      if (!records.length) {
        recordTable.innerHTML = '<tr><td colspan="4"><span class="chip gray">暂无发送记录。</span></td></tr>';
        return;
      }
      recordTable.innerHTML = records.map(row => `
        <tr>
          <td><strong>${esc(row.record_name)}</strong><small>${esc(row.source || "-")}</small></td>
          <td>${esc(row.recipient_count || 0)} 人<small class="mono">${esc((row.recipients || []).slice(0, 3).join(", "))}${(row.recipients || []).length > 3 ? "..." : ""}</small></td>
          <td>${esc(row.last_content || "-")}</td>
          <td class="mono">${esc(row.saved_at || "-")}</td>
        </tr>
      `).join("");
    }

    function renderLogs(logs) {
      if (!logs.length) {
        logTable.innerHTML = '<tr><td colspan="3"><span class="chip gray">暂无发送日志。</span></td></tr>';
        return;
      }
      logTable.innerHTML = logs.map(row => `
        <tr>
          <td class="mono">${esc(row.ts || row.saved_at || "-")}</td>
          <td>${esc(row.template_name || row.record_name || "-")}</td>
          <td>${row.ok ? '<span class="ok">成功</span>' : '<span class="err">失败</span>'}<small>${esc(row.open_id || row.content || "")}</small></td>
        </tr>
      `).join("");
    }

    async function loadState() {
      const data = await api(stateUrl);
      if (!data) return;
      const user = data.current_user || {};
      userInfo.innerHTML = `当前登录：<strong>${esc(user.name || user.open_id || "管理员")}</strong> <span class="mono">(${esc(user.open_id || "-")})</span>`;
      renderStats(data.summary || {});
      renderTags(data.tags || []);
      renderAuth(data.authorized_users || []);
      renderRecords(data.send_records || []);
      renderLogs(data.recent_send_logs || []);
      setStatus("数据已刷新。");
    }

    document.getElementById("refreshBtn").addEventListener("click", () => loadState().catch(err => setStatus(err.message, "err")));
    document.getElementById("logoutBtn").addEventListener("click", () => {
      window.location.href = "/manage/logout";
    });
    document.getElementById("syncBtn").addEventListener("click", async () => {
      try {
        setStatus("正在同步组织成员...");
        const data = await api(syncUrl, {method: "POST", body: "{}"});
        if (!data) return;
        setStatus(data.message || `同步成功：新增 ${data.added}，当前可用 ${data.total}`);
        await loadState();
      } catch (err) {
        setStatus(err.message, "err");
      }
    });
    document.getElementById("restartBtn").addEventListener("click", async () => {
      if (!confirm("确定要重启服务吗？")) return;
      try {
        setStatus("正在触发服务重启...");
        const data = await api(restartUrl, {method: "POST", body: "{}"});
        if (!data) return;
        setStatus(data.message || "已触发重启，请稍后刷新页面。");
      } catch (err) {
        setStatus(err.message, "err");
      }
    });

    loadState().catch(err => setStatus(err.message, "err"));
  </script>
</body>
</html>
    """


@app.get("/manage/login")
def manage_login() -> RedirectResponse:
    if not SETTINGS.public_base_url:
        raise HTTPException(status_code=400, detail="缺少 PUBLIC_BASE_URL，请先配置固定域名")
    if not SETTINGS.app_id or not SETTINGS.app_secret:
        raise HTTPException(status_code=400, detail="缺少 FEISHU_APP_ID / FEISHU_APP_SECRET")
    return RedirectResponse(url=_oauth_login_url(purpose="manage", next_path="/manage"), status_code=302)


@app.get("/manage/logout")
def manage_logout() -> RedirectResponse:
    response = RedirectResponse(url="/manage/login", status_code=302)
    response.delete_cookie(MANAGE_SESSION_COOKIE, path="/")
    return response


@app.get("/manage")
def manage_page(request: Request) -> HTMLResponse:
    try:
        _manage_open_id_from_request(request)
    except HTTPException:
        return RedirectResponse(url="/manage/login", status_code=302)
    return HTMLResponse(_render_manage_page(), media_type="text/html")


@app.get("/admin")
def manage_alias() -> RedirectResponse:
    return RedirectResponse(url="/manage", status_code=302)


@app.get("/manage/api/state")
def manage_api_state(request: Request) -> dict[str, Any]:
    open_id = _manage_open_id_from_request(request)
    return _build_manage_state(open_id)


@app.post("/manage/api/sync-org-contacts")
def manage_api_sync_org_contacts(request: Request) -> dict[str, Any]:
    _manage_open_id_from_request(request)
    added, total, err = _sync_org_contacts()
    if err:
        raise HTTPException(status_code=400, detail=err)
    return {"ok": True, "added": added, "total": total, "message": f"组织成员同步完成：新增{added}，当前可用{total}"}


@app.post("/manage/api/restart-service")
def manage_api_restart_service(request: Request) -> dict[str, Any]:
    _manage_open_id_from_request(request)
    _restart_service_async()
    return {"ok": True, "message": f"已触发 {MANAGE_SERVICE_NAME} 重启，稍后页面会自动恢复。"}


@app.post("/feishu/events")
def feishu_events(payload: EventEnvelope) -> dict[str, Any]:
    if payload.type == "url_verification" and payload.challenge:
        if SETTINGS.verify_token and payload.token != SETTINGS.verify_token:
            raise HTTPException(status_code=401, detail="verify_token 不匹配")
        return {"challenge": payload.challenge}

    header = payload.header or {}
    event_id = header.get("event_id", "")
    if not _consume_event_once(event_id):
        return {"ok": True, "ignored": "duplicate_event"}

    event_type = header.get("event_type")
    if event_type != "im.message.receive_v1":
        return {"ok": True, "ignored": event_type}

    event = payload.event or {}
    message = event.get("message", {})
    sender = event.get("sender", {})

    chat_id = message.get("chat_id", "")
    message_id = message.get("message_id", "")
    msg_type = (message.get("msg_type") or "").strip()
    sender_open_id = sender.get("sender_id", {}).get("open_id", "")
    raw_content = message.get("content") or ""
    content_data = _parse_message_content(raw_content)
    file_key = (
        content_data.get("file_key")
        or content_data.get("fileKey")
        or content_data.get("key")
        or ""
    )
    file_name = (
        content_data.get("file_name")
        or content_data.get("fileName")
        or content_data.get("name")
        or "名单文件"
    )
    is_uploaded_file = bool(file_key and (msg_type == "file" or "file_name" in content_data or "fileName" in content_data))
    text = (content_data.get("text") or "").strip() if content_data else ""
    if not text and isinstance(raw_content, str):
        if not is_uploaded_file:
            text = raw_content.strip()
    text = _norm_text(text) if text else ""
    text = _strip_bot_mention_prefix(text) if text else ""

    if is_uploaded_file:
        if not _is_authorized_user(sender_open_id):
            ret = _reply_chat(chat_id, "你当前没有管理权限，暂不能使用名单文件导入功能。")
            return {"ok": True, "mode": "file_import_denied", "send_result": ret}
        file_blob, download_err = _download_message_resource(message_id, file_key, resource_type="file")
        if not file_blob:
            ret = _reply_chat(chat_id, f"名单文件下载失败：{download_err or '未知错误'}")
            return {"ok": True, "mode": "file_import_download_failed", "send_result": ret}

        names, parse_err = _parse_uploaded_name_file(file_name, file_blob)
        if not names:
            err_text = parse_err or "未识别出姓名，请确认文件为 txt/csv/xlsx，且姓名在单独一列或单独一行。"
            ret = _reply_chat(chat_id, f"已收到文件[{file_name}]，但{err_text}")
            return {"ok": True, "mode": "file_import_parse_failed", "send_result": ret}

        _save_last_uploaded_name_file(
            sender_open_id,
            chat_id=chat_id,
            file_name=file_name,
            names=names,
            message_id=message_id,
            file_key=file_key,
        )
        preview = "、".join(names[:8])
        suffix = "..." if len(names) > 8 else ""
        ret = _reply_chat(
            chat_id,
            f"已收到名单文件[{file_name}]，识别到 {len(names)} 个姓名：{preview}{suffix}\n"
            "下一步请发送：文件建标签 自考",
        )
        _append_debug(
            {
                "kind": "uploaded_name_file",
                "event_id": event_id,
                "chat_id": chat_id,
                "sender_open_id": sender_open_id,
                "file_name": file_name,
                "file_key": file_key,
                "name_count": len(names),
                "msg_type": msg_type,
                "content_keys": sorted(list(content_data.keys())),
            }
        )
        return {"ok": True, "mode": "file_import_uploaded", "send_result": ret, "name_count": len(names)}

    if not text:
        return {"ok": True, "ignored": "empty_text"}
    _append_debug(
        {
            "kind": "incoming_text",
            "event_id": event_id,
            "event_type": event_type,
            "chat_id": chat_id,
            "sender_open_id": sender_open_id,
            "chat_type": message.get("chat_type", ""),
            "mentions_count": len(message.get("mentions", []) or []),
            "text": text,
        }
    )

    # ---------- 简化版群发：第一次发送即保存记录 ----------
    # 用法A（你自己选人，不需要@）：新建群发 晚课提醒 张三,李四 | 今天晚上七点考试
    # 用法B（在群里@学生）：新建群发 晚课提醒 今天晚上七点考试
    if not _consume_recent_command_once(sender_open_id, chat_id, text):
        _append_debug(
            {
                "kind": "duplicate_command_ignored",
                "event_id": event_id,
                "chat_id": chat_id,
                "sender_open_id": sender_open_id,
                "text": text,
            }
        )
        ret = _reply_chat(chat_id, "检测到重复命令：8秒内相同文本会自动去重，请稍后重试。")
        return {"ok": True, "ignored": "duplicate_command", "send_result": ret}

    if _is_super_admin_command(text) and not _is_super_admin(sender_open_id):
        teacher_oid = (SETTINGS.teacher_open_id or "").strip()
        if not _is_valid_open_id(teacher_oid):
            ret = _reply_chat(
                chat_id,
                "你暂无权限管理可用人员。当前主管理员配置异常（TEACHER_OPEN_ID 无效），请联系维护人员在 .env 中改为完整 ou_ 开头ID（不能带...）。",
            )
        else:
            ret = _reply_chat(chat_id, "你暂无权限管理可用人员。只有主管理员可以授权或移除使用权限。")
        return {"ok": True, "mode": "super_admin_denied", "send_result": ret}

    if _is_manager_command(text) and not _is_authorized_user(sender_open_id):
        ret = _reply_chat(
            chat_id,
            "你当前没有管理权限，暂时不能使用群发、标签维护、群绑定这类功能。\n如需开通，请联系管理员授权。",
        )
        return {"ok": True, "mode": "manager_denied", "send_result": ret}

    if text.startswith("新建群发 "):
        body = text.replace("新建群发 ", "", 1).strip()
        recipients: list[str] = []
        roster = _load_roster()

        if "|" in body:
            left, content = body.split("|", 1)
            left = left.strip()
            content = content.strip()
            seg = left.split(maxsplit=1)
            if len(seg) != 2:
                ret = _reply_chat(chat_id, "格式错误：新建群发 记录名 张三,李四 | 文案")
                return {"ok": True, "mode": "quick_new_send", "send_result": ret}
            record_name = seg[0].strip()
            names = _split_csv_names(seg[1])
            if not names:
                ret = _reply_chat(chat_id, "未识别到学生名单，请用逗号分隔姓名。")
                return {"ok": True, "mode": "quick_new_send", "send_result": ret}
            missing: list[str] = []
            for n in names:
                oid = roster.get(n)
                if oid:
                    recipients.append(oid)
                else:
                    missing.append(n)
            if not recipients:
                ret = _reply_chat(chat_id, f"发送失败：名单里都未绑定。未绑定：{','.join(missing)}")
                return {"ok": True, "mode": "quick_new_send", "send_result": ret}
        else:
            seg = body.split(maxsplit=1)
            if len(seg) != 2:
                ret = _reply_chat(chat_id, "格式错误：新建群发 记录名 文案（并@学生）")
                return {"ok": True, "mode": "quick_new_send", "send_result": ret}
            record_name, content = seg[0].strip(), seg[1].strip()
            mention_rows = _extract_mentions(message.get("mentions", []))
            if not mention_rows:
                ret = _reply_chat(chat_id, "请先@要发送的学生，或使用：新建群发 记录名 张三,李四 | 文案")
                return {"ok": True, "mode": "quick_new_send", "send_result": ret}
            for name, oid in mention_rows:
                if name:
                    roster[name] = oid
                recipients.append(oid)
            _save_roster(roster)

        recipients = _dedup_keep_order(recipients)
        if sender_open_id and recipients and all(rid == sender_open_id for rid in recipients):
            ret = _reply_chat(
                chat_id,
                "本次名单只匹配到了你自己，所以没有发给其他学员。\n"
                "请先让学员发送：学员注册 姓名，或先“同步组织成员”后用“搜索成员 姓名”确认已入库。",
            )
            return {"ok": True, "mode": "quick_new_send", "send_result": ret}

        logs = _read_json(SEND_LOG_FILE, [])
        success = 0
        for rid in recipients:
            ret_send = _send_text(rid, content, receive_id_type="open_id")
            ok = ret_send.get("code") == 0
            if ok:
                success += 1
            logs.append(
                {
                    "record_name": record_name,
                    "open_or_chat_id": rid,
                    "content": content,
                    "ok": ok,
                    "raw": ret_send,
                }
            )
        _write_json(SEND_LOG_FILE, logs)

        records = _load_send_records()
        records[record_name] = {
            "owner": sender_open_id,
            "recipients": recipients,
            "last_content": content,
            "source": "quick_new_send",
            "saved_at": datetime.now().isoformat(),
        }
        _save_send_records(records)
        _save_sender_last_context(sender_open_id=sender_open_id, recipients=recipients, content=content, source=f"record:{record_name}")

        ret = _reply_chat(chat_id, f"新建并发送完成：{record_name}，成功 {success}/{len(recipients)}。下次可用：再发群发 {record_name} 新文案")
        return {"ok": True, "mode": "quick_new_send", "send_result": ret}

    # ---------- 简化版群发：按记录再发 ----------
    # 用法：再发群发 晚课提醒 明晚改到19:30
    if text.startswith("再发群发 "):
        body = text.replace("再发群发 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误：再发群发 记录名 新文案")
            return {"ok": True, "mode": "quick_resend", "send_result": ret}
        record_name, content = seg[0].strip(), seg[1].strip()
        records = _load_send_records()
        record = records.get(record_name)
        if not record:
            ret = _reply_chat(chat_id, f"未找到群发记录：{record_name}")
            return {"ok": True, "mode": "quick_resend", "send_result": ret}

        recipients = record.get("recipients", [])
        if not recipients:
            ret = _reply_chat(chat_id, f"群发记录[{record_name}]没有接收人")
            return {"ok": True, "mode": "quick_resend", "send_result": ret}

        logs = _read_json(SEND_LOG_FILE, [])
        success = 0
        for rid in recipients:
            ret_send = _send_text(rid, content, receive_id_type="open_id" if str(rid).startswith("ou_") else "chat_id")
            ok = ret_send.get("code") == 0
            if ok:
                success += 1
            logs.append(
                {
                    "record_name": record_name,
                    "open_or_chat_id": rid,
                    "content": content,
                    "ok": ok,
                    "raw": ret_send,
                }
            )
        _write_json(SEND_LOG_FILE, logs)
        _save_sender_last_context(sender_open_id=sender_open_id, recipients=recipients, content=content, source=f"record:{record_name}")
        ret = _reply_chat(chat_id, f"再发完成：{record_name}，成功 {success}/{len(recipients)}")
        return {"ok": True, "mode": "quick_resend", "send_result": ret}

    if text in {"帮助", "幫助", "help", "菜单", "命令", "甯姪", "骞姪", "鑿滃崟", "鍛戒护"}:
        help_text = (
            "常用命令：\n"
            "1) 查看标签 自考\n"
            "2) 查看可用人员\n"
            "3) 查看成员 雷炫 / 查看人员 雷炫\n"
            "4) 管理页 / 打开管理页\n"
            "5) 我的身份 / 版本\n"
            "6) 上传名单文件后：文件建标签 自考\n"
            "7) 复用群发 自考 今晚20:00上课提醒\n"
            "8) 群标签群发 A组 今晚20:00上课提醒\n"
            "9) 授权使用 张三\n"
            "10) 移除使用权限 张三\n"
            "11) 授权接收预警 张三 / 移除预警接收 张三\n"
            "\n"
            "说明：\n"
            "先把 txt/csv/xlsx 名单文件发给机器人，再发：文件建标签 自考\n"
            "文件建标签会优先按组织通讯录自动匹配可发送人员。\n"
            "查看成员/查看人员 雷炫：显示该成员完整 open_id，方便你做授权。\n"
            "复用群发格式：复用群发 自考 今晚20:00上课提醒\n"
            "复用群发会优先按模板名发送；如果没有同名模板，就按学生标签发送。\n"
            "主管理员权限默认保留，不会因授权他人而丢失。\n"
            "授权使用/移除使用权限优先按姓名匹配；如果重名，再用：授权使用 张三 ou_xxx\n"
            "授权接收预警/移除预警接收优先按姓名匹配；如果重名，再用：授权接收预警 张三 ou_xxx"
        )
        ret = _reply_chat(chat_id, help_text)
        return {"ok": True, "mode": "help", "send_result": ret}


    if text in {"管理页", "打开管理页", "后台"}:
        url = _manage_page_url()
        if not url:
            ret = _reply_chat(chat_id, "当前还没有配置 PUBLIC_BASE_URL，无法生成管理页链接。")
            return {"ok": True, "mode": "manage_page", "send_result": ret}
        ret = _reply_chat(chat_id, f"管理页地址：{url}\n打开后会先进行飞书登录，登录完成即可查看标签、授权人员和发送记录。")
        return {"ok": True, "mode": "manage_page", "send_result": ret}

    if text in {"版本", "ver", "version"}:
        teacher_oid = (SETTINGS.teacher_open_id or "").strip()
        teacher_show = teacher_oid if not teacher_oid else f"{teacher_oid[:12]}...{teacher_oid[-6:]}"
        ret = _reply_chat(
            chat_id,
            f"当前版本：{BOT_BUILD}\n"
            f"主管理员ID：{teacher_show}\n"
            "如果你刚更新过代码但版本没变，说明服务器还在旧代码。",
        )
        return {"ok": True, "mode": "version", "send_result": ret}

    if text in {"我的身份", "我的权限", "whoami"}:
        is_super = _is_super_admin(sender_open_id)
        is_auth = _is_authorized_user(sender_open_id)
        ret = _reply_chat(
            chat_id,
            f"你的 open_id：{sender_open_id}\n"
            f"主管理员：{'是' if is_super else '否'}\n"
            f"管理权限：{'有' if is_auth else '无'}",
        )
        return {"ok": True, "mode": "whoami", "send_result": ret}


    if text in {"查看可用人员", "查看授权人员", "查看使用权限"}:
        rows = _load_authorized_users()
        teacher_oid = (SETTINGS.teacher_open_id or "").strip()
        lines = []
        for open_id, name in rows.items():
            suffix = "（主管理员）" if teacher_oid and open_id == teacher_oid else ""
            lines.append(f"{name or open_id} -> {open_id[:12]}...{suffix}")
        if not lines:
            ret = _reply_chat(chat_id, "当前还没有可用人员。请先检查 .env 中 TEACHER_OPEN_ID 是否配置正确。")
            return {"ok": True, "mode": "show_authorized_users", "send_result": ret}
        ret = _reply_chat(chat_id, "当前可用人员：\n" + "\n".join(lines))
        return {"ok": True, "mode": "show_authorized_users", "send_result": ret}

    if text.startswith("查看成员 ") or text.startswith("查看人员 "):
        body = text.replace("查看成员 ", "", 1).replace("查看人员 ", "", 1).strip()
        if not body:
            ret = _reply_chat(chat_id, "格式错误，请使用：查看成员 雷炫 或 查看人员 雷炫")
            return {"ok": True, "mode": "show_member", "send_result": ret}
        candidates = _find_member_candidates_by_name(body)
        if not candidates:
            ret = _reply_chat(chat_id, f"未找到成员：{body}\n可以先发：同步组织成员")
            return {"ok": True, "mode": "show_member", "send_result": ret}
        lines = []
        for idx, item in enumerate(candidates, start=1):
            source = item.get("source", "")
            suffix = f"（{source}）" if source else ""
            lines.append(f"{idx}) {item['name']} -> {item['open_id']}{suffix}")
        ret = _reply_chat(chat_id, "成员查询结果：\n" + "\n".join(lines))
        return {"ok": True, "mode": "show_member", "send_result": ret}

    if text in {"查看预警接收人", "查看预警权限"}:
        rows = _load_alert_receivers()
        teacher_oid = (SETTINGS.teacher_open_id or "").strip()
        teacher_name = _resolve_name_by_open_id(teacher_oid) or "主管理员"
        lines = []
        if _is_valid_open_id(teacher_oid):
            lines.append(f"{teacher_name} -> {teacher_oid[:12]}...（默认主接收人）")
        for open_id, name in rows.items():
            lines.append(f"{name or open_id} -> {open_id[:12]}...")
        if not lines:
            ret = _reply_chat(chat_id, "当前还没有配置预警接收人。")
            return {"ok": True, "mode": "show_alert_receivers", "send_result": ret}
        ret = _reply_chat(chat_id, "当前预警接收人：\n" + "\n".join(lines))
        return {"ok": True, "mode": "show_alert_receivers", "send_result": ret}

    if text.startswith("授权使用 "):
        body = text.replace("授权使用 ", "", 1).strip()
        target_open_id, display_name = _resolve_authorize_target(body, message.get("mentions", []))
        if not _is_valid_open_id(target_open_id):
            ret = _reply_chat(chat_id, "格式错误，请使用：授权使用 张三 ou_xxx，或在群里@对方后发送：授权使用 张三")
            return {"ok": True, "mode": "grant_authorized_user", "send_result": ret}
        _grant_authorized_user(target_open_id, display_name)
        final_name = display_name or _resolve_name_by_open_id(target_open_id) or target_open_id
        ret = _reply_chat(chat_id, f"已授权使用：{final_name}")
        return {"ok": True, "mode": "grant_authorized_user", "send_result": ret}

    if text.startswith("移除使用权限 "):
        body = text.replace("移除使用权限 ", "", 1).strip()
        target_open_id, display_name = _resolve_authorize_target(body, message.get("mentions", []))
        if not _is_valid_open_id(target_open_id):
            ret = _reply_chat(chat_id, "格式错误，请使用：移除使用权限 张三 ou_xxx，或在群里@对方后发送：移除使用权限 张三")
            return {"ok": True, "mode": "revoke_authorized_user", "send_result": ret}
        if _is_super_admin(target_open_id):
            ret = _reply_chat(chat_id, "主管理员权限不可移除。")
            return {"ok": True, "mode": "revoke_authorized_user", "send_result": ret}
        ok = _revoke_authorized_user(target_open_id)
        if not ok:
            ret = _reply_chat(chat_id, "该人员当前不在授权名单中。")
            return {"ok": True, "mode": "revoke_authorized_user", "send_result": ret}
        final_name = display_name or _resolve_name_by_open_id(target_open_id) or target_open_id
        ret = _reply_chat(chat_id, f"已移除使用权限：{final_name}")
        return {"ok": True, "mode": "revoke_authorized_user", "send_result": ret}

    if text.startswith("授权接收预警 "):
        body = text.replace("授权接收预警 ", "", 1).strip()
        target_open_id, display_name = _resolve_authorize_target(body, message.get("mentions", []))
        if not _is_valid_open_id(target_open_id):
            ret = _reply_chat(chat_id, "格式错误，请使用：授权接收预警 张三 ou_xxx，或在群里@对方后发送：授权接收预警 张三")
            return {"ok": True, "mode": "grant_alert_receiver", "send_result": ret}
        _grant_alert_receiver(target_open_id, display_name)
        final_name = display_name or _resolve_name_by_open_id(target_open_id) or target_open_id
        ret = _reply_chat(chat_id, f"已授权接收预警：{final_name}")
        return {"ok": True, "mode": "grant_alert_receiver", "send_result": ret}

    if text.startswith("移除预警接收 "):
        body = text.replace("移除预警接收 ", "", 1).strip()
        target_open_id, display_name = _resolve_authorize_target(body, message.get("mentions", []))
        if not _is_valid_open_id(target_open_id):
            ret = _reply_chat(chat_id, "格式错误，请使用：移除预警接收 张三 ou_xxx，或在群里@对方后发送：移除预警接收 张三")
            return {"ok": True, "mode": "revoke_alert_receiver", "send_result": ret}
        ok = _revoke_alert_receiver(target_open_id)
        if not ok:
            ret = _reply_chat(chat_id, "该人员当前不在预警接收名单中。")
            return {"ok": True, "mode": "revoke_alert_receiver", "send_result": ret}
        final_name = display_name or _resolve_name_by_open_id(target_open_id) or target_open_id
        ret = _reply_chat(chat_id, f"已移除预警接收：{final_name}")
        return {"ok": True, "mode": "revoke_alert_receiver", "send_result": ret}

    lower_text = text.lower()
    if (
        text in {"同步组织成员", "同步成员", "同步通讯录"}
        or ("同步" in text and ("组织" in text or "成员" in text or "通讯录" in text))
        or lower_text in {"sync", "sync_org", "sync_contacts"}
    ):
        added, total, err = _sync_org_contacts()
        if err:
            ret = _reply_chat(chat_id, err)
            return {"ok": True, "mode": "sync_org_contacts", "send_result": ret}
        ret = _reply_chat(chat_id, f"组织成员同步完成：新增{added}，当前可用{total}")
        return {"ok": True, "mode": "sync_org_contacts", "send_result": ret}

    if text.startswith("搜索成员 "):
        kw = text.replace("搜索成员 ", "", 1).strip()
        if not kw:
            ret = _reply_chat(chat_id, "格式错误，请使用：搜索成员 张")
            return {"ok": True, "mode": "search_members", "send_result": ret}
        roster = _load_roster()
        matches = [name for name in roster.keys() if kw in name][:20]
        if not matches:
            ret = _reply_chat(chat_id, f"未找到包含“{kw}”的成员。先试试：同步组织成员")
            return {"ok": True, "mode": "search_members", "send_result": ret}
        ret = _reply_chat(chat_id, "匹配成员：\n" + "\n".join(matches))
        return {"ok": True, "mode": "search_members", "send_result": ret}

    if text == "查看绑定":
        roster = _load_roster()
        if not roster:
            ret = _reply_chat(chat_id, "当前没有任何学员绑定。")
            return {"ok": True, "mode": "show_roster", "send_result": ret}
        lines = [f"{name} -> {oid[:10]}..." for name, oid in list(roster.items())[:20]]
        ret = _reply_chat(chat_id, "当前绑定：\n" + "\n".join(lines))
        return {"ok": True, "mode": "show_roster", "send_result": ret}

    if text == "查看群标签":
        tags = _load_chat_tags()
        if not tags:
            ret = _reply_chat(chat_id, "当前还没有保存任何群标签。")
            return {"ok": True, "mode": "show_chat_tags", "send_result": ret}
        lines = [f"{name} -> {cid[:12]}..." for name, cid in list(tags.items())[:20]]
        ret = _reply_chat(chat_id, "当前群标签：\n" + "\n".join(lines))
        return {"ok": True, "mode": "show_chat_tags", "send_result": ret}

    if text.startswith("查看名单 "):
        template_name = text.replace("查看名单 ", "", 1).strip()
        rows = _read_json(TEMPLATE_FILE, [])
        row = next((r for r in rows if r.get("template_name") == template_name), None)
        if not row:
            ret = _reply_chat(chat_id, f"未找到名单模板：{template_name}")
            return {"ok": True, "mode": "show_template", "send_result": ret}
        names = row.get("recipient_names", [])
        ids = row.get("recipient_open_ids", [])
        ret = _reply_chat(chat_id, f"名单[{template_name}]：姓名{len(names)}人，有效接收人{len(ids)}人")
        return {"ok": True, "mode": "show_template", "send_result": ret}

    # ---------- 命令：学员自注册（无需老师输入 open_id） ----------
    # Format: 学员注册 张三
    if text.startswith("查看标签 "):
        tag_name = text.replace("查看标签 ", "", 1).strip()
        names = _get_students_by_tag(tag_name)
        if not names:
            ret = _reply_chat(chat_id, f"标签[{tag_name}]下还没有学员。")
            return {"ok": True, "mode": "show_student_tag", "send_result": ret}
        ret = _reply_chat(chat_id, f"标签[{tag_name}]名单：\n" + "\n".join(names))
        return {"ok": True, "mode": "show_student_tag", "send_result": ret}

    if text in {"查看所有标签", "查看标签列表", "标签列表"}:
        tag_names = _list_all_tag_names()
        if not tag_names:
            ret = _reply_chat(chat_id, "当前还没有任何标签。")
            return {"ok": True, "mode": "list_student_tags", "send_result": ret}
        ret = _reply_chat(chat_id, "当前标签：\n" + "\n".join(tag_names))
        return {"ok": True, "mode": "list_student_tags", "send_result": ret}

    if text.startswith("文件建标签 "):
        tag_name = text.replace("文件建标签 ", "", 1).strip()
        if not tag_name:
            ret = _reply_chat(chat_id, "格式错误，请使用：文件建标签 自考")
            return {"ok": True, "mode": "file_build_tag", "send_result": ret}
        file_ctx = _get_last_uploaded_name_file(sender_open_id)
        file_name = str(file_ctx.get("file_name", "")).strip()
        names = file_ctx.get("names", [])
        if not file_name or not isinstance(names, list) or not names:
            ret = _reply_chat(chat_id, "你还没有上传可用名单文件，请先把 txt/csv/xlsx 文件发给机器人。")
            return {"ok": True, "mode": "file_build_tag", "send_result": ret}
        final_names = _replace_students_in_tag(tag_name, [str(x).strip() for x in names if str(x).strip()])
        matched_names, missing_names, sync_added, sync_error = _hydrate_roster_by_names(final_names, auto_sync_if_empty=True)
        if missing_names and not SETTINGS.mock_send:
            refresh_added, _, refresh_error = _sync_org_contacts()
            if refresh_added:
                sync_added += refresh_added
            if refresh_error:
                sync_error = refresh_error
            matched_names, missing_names, _, _ = _hydrate_roster_by_names(final_names, auto_sync_if_empty=False)
        preview = "、".join(final_names[:10])
        suffix = "..." if len(final_names) > 10 else ""
        extra_lines: list[str] = []
        if sync_added:
            extra_lines.append(f"已自动同步组织成员，新增 {sync_added} 人。")
        extra_lines.append(f"当前可直接发送 {len(matched_names)}/{len(final_names)} 人。")
        if missing_names:
            extra_lines.append(f"仍未匹配：{'、'.join(missing_names[:10])}{'...' if len(missing_names) > 10 else ''}")
        if sync_error:
            extra_lines.append(f"通讯录自动同步未完成：{sync_error}")
        ret = _reply_chat(
            chat_id,
            f"已按文件[{file_name}]更新标签[{tag_name}]，共 {len(final_names)} 人：{preview}{suffix}\n"
            + "\n".join(extra_lines)
            + f"\n现在可直接发送：复用群发 {tag_name} 今晚20:00上课提醒",
        )
        return {"ok": True, "mode": "file_build_tag", "send_result": ret}

    if text.startswith("添加到标签 "):
        body = text.replace("添加到标签 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：添加到标签 自考 张三,李四")
            return {"ok": True, "mode": "add_students_to_tag", "send_result": ret}
        tag_name = seg[0].strip()
        names = _split_csv_names(seg[1].replace("，", ","))
        if not names:
            ret = _reply_chat(chat_id, "请至少填写一个学员姓名。")
            return {"ok": True, "mode": "add_students_to_tag", "send_result": ret}
        final_names = _add_students_to_tag(tag_name, names)
        ret = _reply_chat(chat_id, f"已加入标签[{tag_name}]：{', '.join(names)}\n当前名单：{', '.join(final_names)}")
        return {"ok": True, "mode": "add_students_to_tag", "send_result": ret}

    if text.startswith("从标签移除 "):
        body = text.replace("从标签移除 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：从标签移除 自考 张三")
            return {"ok": True, "mode": "remove_students_from_tag", "send_result": ret}
        tag_name = seg[0].strip()
        names = _split_csv_names(seg[1].replace("，", ","))
        if not names:
            ret = _reply_chat(chat_id, "请至少填写一个学员姓名。")
            return {"ok": True, "mode": "remove_students_from_tag", "send_result": ret}
        final_names, existed = _remove_students_from_tag(tag_name, names)
        if not existed:
            ret = _reply_chat(chat_id, f"标签[{tag_name}]不存在。")
            return {"ok": True, "mode": "remove_students_from_tag", "send_result": ret}
        if final_names:
            ret = _reply_chat(chat_id, f"已从标签[{tag_name}]移除：{', '.join(names)}\n当前名单：{', '.join(final_names)}")
        else:
            ret = _reply_chat(chat_id, f"已从标签[{tag_name}]移除：{', '.join(names)}\n该标签已为空，系统已自动删除。")
        return {"ok": True, "mode": "remove_students_from_tag", "send_result": ret}

    if text.startswith("重命名标签 "):
        body = text.replace("重命名标签 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：重命名标签 自考 成考")
            return {"ok": True, "mode": "rename_student_tag", "send_result": ret}
        ok, err = _rename_student_tag(seg[0].strip(), seg[1].strip())
        if not ok:
            ret = _reply_chat(chat_id, err)
            return {"ok": True, "mode": "rename_student_tag", "send_result": ret}
        ret = _reply_chat(chat_id, f"标签已重命名：{seg[0].strip()} -> {seg[1].strip()}")
        return {"ok": True, "mode": "rename_student_tag", "send_result": ret}

    if text.startswith("删除标签 "):
        tag_name = text.replace("删除标签 ", "", 1).strip()
        if not tag_name:
            ret = _reply_chat(chat_id, "格式错误，请使用：删除标签 自考")
            return {"ok": True, "mode": "delete_student_tag", "send_result": ret}
        ok = _delete_student_tag(tag_name)
        if not ok:
            ret = _reply_chat(chat_id, f"标签[{tag_name}]不存在。")
            return {"ok": True, "mode": "delete_student_tag", "send_result": ret}
        ret = _reply_chat(chat_id, f"已删除标签：{tag_name}")
        return {"ok": True, "mode": "delete_student_tag", "send_result": ret}

    if text.startswith("学员注册 "):
        body = text.replace("学员注册 ", "", 1).strip()
        if not body:
            ret = _reply_chat(chat_id, "格式错误，请使用：学员注册 张三 或 学员注册 自考 张三")
            return {"ok": True, "mode": "self_register", "send_result": ret}
        parts = body.split()
        tag_names: list[str] = []
        if len(parts) == 1:
            name = parts[0].strip()
        else:
            tag_names = _split_tag_names(parts[0].strip())
            name = " ".join(parts[1:]).strip()
        if not name:
            ret = _reply_chat(chat_id, "格式错误，请使用：学员注册 张三 或 学员注册 自考 张三")
            return {"ok": True, "mode": "self_register", "send_result": ret}
        if not sender_open_id:
            ret = _reply_chat(chat_id, "注册失败：未识别发送者 open_id")
            return {"ok": True, "mode": "self_register", "send_result": ret}
        roster = _load_roster()
        roster[name] = sender_open_id
        _save_roster(roster)
        saved_tags = _bind_tags_for_student(name, tag_names) if tag_names else []
        msg = f"注册成功：{name}"
        if saved_tags:
            msg += f"\n已加入标签：{', '.join(saved_tags)}"
        ret = _reply_chat(chat_id, msg)
        return {"ok": True, "mode": "self_register", "send_result": ret}

    if text.startswith("按标签群发 "):
        body = text.replace("按标签群发 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：按标签群发 自考 今晚20:00上课提醒")
            return {"ok": True, "mode": "tag_broadcast", "send_result": ret}
        tag_name = seg[0].strip()
        content = seg[1].strip()
        result = _send_with_tag(tag_name, content)
        if not result.get("ok"):
            ret = _reply_chat(chat_id, f"发送失败：{result.get('msg')}")
            return {"ok": True, "mode": "tag_broadcast", "send_result": ret}
        recipients = [item["open_id"] for item in result.get("results", [])]
        _save_sender_last_context(sender_open_id=sender_open_id, recipients=recipients, content=content, source=f"tag:{tag_name}")
        missing = result.get("missing_names", [])
        if missing:
            msg = f"按标签群发完成：[{tag_name}]，成功 {result['success']}/{result['total']}，未绑定：{','.join(missing)}"
        else:
            msg = f"按标签群发完成：[{tag_name}]，成功 {result['success']}/{result['total']}"
        ret = _reply_chat(chat_id, msg)
        return {"ok": True, "mode": "tag_broadcast", "send_result": ret}

    if text.startswith("学员注册 "):
        name = text.replace("学员注册 ", "", 1).strip()
        if not name:
            ret = _reply_chat(chat_id, "格式错误，请使用：学员注册 张三")
            return {"ok": True, "mode": "self_register", "send_result": ret}
        if not sender_open_id:
            ret = _reply_chat(chat_id, "注册失败：未识别发送者 open_id")
            return {"ok": True, "mode": "self_register", "send_result": ret}
        roster = _load_roster()
        roster[name] = sender_open_id
        _save_roster(roster)
        ret = _reply_chat(chat_id, f"注册成功：{name}")
        return {"ok": True, "mode": "self_register", "send_result": ret}

    # ---------- 命令：设置预警阈值 ----------
    if text.startswith("设置预警阈值 "):
        level = text.replace("设置预警阈值 ", "", 1).strip()
        if level not in {"低", "中", "高"}:
            ret = _reply_chat(chat_id, "阈值仅支持：低 / 中 / 高")
            return {"ok": True, "mode": "set_threshold", "send_result": ret}
        SETTINGS.alert_level = level
        ret = _reply_chat(chat_id, f"已设置自动预警阈值：{level}")
        return {"ok": True, "mode": "set_threshold", "send_result": ret}

    # ---------- 命令：绑定学员 ----------
    if text.startswith("绑定学员 "):
        body = text.replace("绑定学员 ", "", 1).strip()
        parts = body.split()
        student_name = ""
        student_open_id = ""
        if len(parts) == 2 and parts[1].startswith("ou_"):
            student_name = parts[0]
            student_open_id = parts[1]
        elif len(parts) == 1:
            # Teacher can @student in group, then bind by name without manually copying open_id.
            student_name = parts[0]
            student_open_id = _extract_open_id_from_mentions(message.get("mentions", []))
        if not student_name or not student_open_id:
            ret = _reply_chat(chat_id, "格式错误，请使用：绑定学员 张三 ou_xxx，或在群里@学生后发送：绑定学员 张三")
            return {"ok": True, "mode": "bind_student", "send_result": ret}
        roster = _load_roster()
        roster[student_name] = student_open_id
        _save_roster(roster)
        ret = _reply_chat(chat_id, f"已绑定学员：{student_name}")
        return {"ok": True, "mode": "bind_student", "send_result": ret}

    # ---------- 命令：保存名单 ----------
    if text.startswith("保存名单 "):
        body = text.replace("保存名单 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：保存名单 A组 张三,李四")
            return {"ok": True, "mode": "save_template", "send_result": ret}
        template_name = seg[0].strip()
        names = _split_csv_names(seg[1])
        if not names:
            ret = _reply_chat(chat_id, "名单为空，请至少填写一个学员姓名。")
            return {"ok": True, "mode": "save_template", "send_result": ret}
        ok_count, total_count, _, missing = _save_template_by_names(template_name, names)
        if missing:
            ret = _reply_chat(chat_id, f"模板已保存：{template_name}。已匹配 {ok_count}/{total_count}，未绑定学员：{','.join(missing)}")
        else:
            ret = _reply_chat(chat_id, f"模板已保存：{template_name}。已匹配 {ok_count}/{total_count}。")
        return {"ok": True, "mode": "save_template", "send_result": ret}

    # ---------- 命令：记住当前群并打标签 ----------
    # Format: 记住本群 A组
    if text.startswith("绑定群 "):
        body = text.replace("绑定群 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：绑定群 A组 群名称关键词")
            return {"ok": True, "mode": "bind_chat_tag_by_name", "send_result": ret}
        tag_name = seg[0].strip()
        keyword = seg[1].strip()
        found_chat_id, found_chat_name, err = _find_chat_by_name(keyword)
        if err:
            ret = _reply_chat(chat_id, err)
            return {"ok": True, "mode": "bind_chat_tag_by_name", "send_result": ret}
        tags = _load_chat_tags()
        tags[tag_name] = found_chat_id
        _save_chat_tags(tags)
        ret = _reply_chat(chat_id, f"绑定成功：{tag_name} -> {found_chat_name}")
        return {"ok": True, "mode": "bind_chat_tag_by_name", "send_result": ret}

    if text.startswith("记住本群 "):
        tag_name = text.replace("记住本群 ", "", 1).strip()
        if not tag_name:
            ret = _reply_chat(chat_id, "格式错误，请使用：记住本群 A组")
            return {"ok": True, "mode": "save_chat_tag", "send_result": ret}
        if not chat_id:
            ret = _reply_chat(chat_id, "当前会话无 chat_id，无法保存群标签。")
            return {"ok": True, "mode": "save_chat_tag", "send_result": ret}
        tags = _load_chat_tags()
        tags[tag_name] = chat_id
        _save_chat_tags(tags)
        ret = _reply_chat(chat_id, f"已记住当前群，标签：{tag_name}")
        return {"ok": True, "mode": "save_chat_tag", "send_result": ret}

    # ---------- 命令：复用群发 ----------
    if text.startswith("复用群发 ") or text.startswith("群发 "):
        body = text.replace("复用群发 ", "", 1) if text.startswith("复用群发 ") else text.replace("群发 ", "", 1)
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：复用群发 A组 通知内容")
            return {"ok": True, "mode": "broadcast", "send_result": ret}
        template_name = seg[0].strip()
        content = seg[1].strip()
        result = _send_with_template(template_name, content)
        source = f"template:{template_name}"
        if not result.get("ok"):
            tag_result = _send_with_tag(template_name, content)
            if not tag_result.get("ok"):
                ret = _reply_chat(chat_id, f"发送失败：{result.get('msg')}")
                return {"ok": True, "mode": "broadcast", "send_result": ret, "result": result}
            result = tag_result
            source = f"tag:{template_name}"
        recipients = [item["open_id"] for item in result.get("results", [])]
        _save_sender_last_context(sender_open_id=sender_open_id, recipients=recipients, content=content, source=source)
        missing = result.get("missing_names", [])
        if source.startswith("tag:") and missing:
            msg = f"群发完成：标签[{template_name}]，成功 {result['success']}/{result['total']}，未绑定：{','.join(missing)}"
        elif source.startswith("tag:"):
            msg = f"群发完成：标签[{template_name}]，成功 {result['success']}/{result['total']}"
        else:
            msg = f"群发完成：模板[{template_name}]，成功 {result['success']}/{result['total']}"
        ret = _reply_chat(chat_id, msg)
        return {"ok": True, "mode": "broadcast", "send_result": ret, "result": result}

    # ---------- 命令：按群标签群发 ----------
    # Format: 群标签群发 A组 文案
    if text.startswith("群标签群发 "):
        body = text.replace("群标签群发 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：群标签群发 A组 通知内容")
            return {"ok": True, "mode": "chat_tag_broadcast", "send_result": ret}
        tag_name = seg[0].strip()
        content = seg[1].strip()
        tags = _load_chat_tags()
        target_chat_id = tags.get(tag_name, "")
        if not target_chat_id:
            tag_names = "、".join(list(tags.keys())[:10]) if tags else "无"
            ret = _reply_chat(chat_id, f"未找到群标签：{tag_name}。当前已有：{tag_names}")
            return {"ok": True, "mode": "chat_tag_broadcast", "send_result": ret}
        send_ret = _send_text(target_chat_id, content, receive_id_type="chat_id")
        # 群标签群发也记录一条“最近一次群发”
        if sender_open_id:
            _save_sender_last_context(sender_open_id=sender_open_id, recipients=[target_chat_id], content=content, source=f"chat_tag:{tag_name}")
        ack = _reply_chat(chat_id, f"已按群标签发送：{tag_name}")
        return {"ok": True, "mode": "chat_tag_broadcast", "send_result": send_ret, "ack_result": ack}

    # ---------- 命令：按群成员逐个私发 ----------
    # Format: 群成员私发 A组 文案
    if text.startswith("群成员私发 ") or text.startswith("按群私发 "):
        body = text.replace("群成员私发 ", "", 1) if text.startswith("群成员私发 ") else text.replace("按群私发 ", "", 1)
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：群成员私发 A组 通知内容")
            return {"ok": True, "mode": "chat_member_private_broadcast", "send_result": ret}
        tag_name = seg[0].strip()
        content = seg[1].strip()
        tags = _load_chat_tags()
        target_chat_id = tags.get(tag_name, "")
        if not target_chat_id:
            tag_names = "、".join(list(tags.keys())[:10]) if tags else "无"
            ret = _reply_chat(chat_id, f"未找到群标签：{tag_name}。当前已有：{tag_names}")
            return {"ok": True, "mode": "chat_member_private_broadcast", "send_result": ret}

        members, err = _list_chat_member_open_ids(target_chat_id)
        if err:
            ret = _reply_chat(chat_id, err)
            return {"ok": True, "mode": "chat_member_private_broadcast", "send_result": ret}
        if sender_open_id:
            members = [x for x in members if x != sender_open_id]
        if not members:
            ret = _reply_chat(chat_id, "没有可发送的群成员（已排除你自己）。")
            return {"ok": True, "mode": "chat_member_private_broadcast", "send_result": ret}

        logs = _read_json(SEND_LOG_FILE, [])
        success = 0
        for open_id in members:
            send_ret = _send_text(open_id, content, receive_id_type="open_id")
            ok = send_ret.get("code") == 0
            if ok:
                success += 1
            logs.append(
                {
                    "chat_tag": tag_name,
                    "open_id": open_id,
                    "content": content,
                    "ok": ok,
                    "raw": send_ret,
                }
            )
        _write_json(SEND_LOG_FILE, logs)
        _save_sender_last_context(sender_open_id=sender_open_id, recipients=members, content=content, source=f"chat_members:{tag_name}")
        ret = _reply_chat(chat_id, f"群成员私发完成：{tag_name}，成功 {success}/{len(members)}")
        return {"ok": True, "mode": "chat_member_private_broadcast", "send_result": ret}

    # ---------- 命令：保存最近一次群发为记录 ----------
    # Format: 保存群发记录 晚课提醒
    if text.startswith("保存群发记录 "):
        record_name = text.replace("保存群发记录 ", "", 1).strip()
        if not record_name:
            ret = _reply_chat(chat_id, "格式错误，请使用：保存群发记录 晚课提醒")
            return {"ok": True, "mode": "save_send_record", "send_result": ret}
        ctx = _load_last_send_ctx().get(sender_open_id, {})
        recipients = ctx.get("recipients", [])
        if not recipients:
            ret = _reply_chat(chat_id, "你还没有可保存的最近群发记录，请先执行一次群发。")
            return {"ok": True, "mode": "save_send_record", "send_result": ret}
        records = _load_send_records()
        records[record_name] = {
            "owner": sender_open_id,
            "recipients": recipients,
            "last_content": ctx.get("content", ""),
            "source": ctx.get("source", ""),
            "saved_at": datetime.now().isoformat(),
        }
        _save_send_records(records)
        ret = _reply_chat(chat_id, f"已保存群发记录：{record_name}（接收人数 {len(recipients)}）")
        return {"ok": True, "mode": "save_send_record", "send_result": ret}

    # ---------- 命令：按记录群发（改文案直接发） ----------
    # Format: 按记录群发 晚课提醒 新内容
    if text.startswith("按记录群发 "):
        body = text.replace("按记录群发 ", "", 1).strip()
        seg = body.split(maxsplit=1)
        if len(seg) != 2:
            ret = _reply_chat(chat_id, "格式错误，请使用：按记录群发 晚课提醒 新文案")
            return {"ok": True, "mode": "send_by_record", "send_result": ret}
        record_name, content = seg[0].strip(), seg[1].strip()
        records = _load_send_records()
        record = records.get(record_name)
        if not record:
            ret = _reply_chat(chat_id, f"未找到群发记录：{record_name}")
            return {"ok": True, "mode": "send_by_record", "send_result": ret}
        recipients = record.get("recipients", [])
        if not recipients:
            ret = _reply_chat(chat_id, f"群发记录[{record_name}]没有接收人")
            return {"ok": True, "mode": "send_by_record", "send_result": ret}

        logs = _read_json(SEND_LOG_FILE, [])
        success = 0
        for rid in recipients:
            if str(rid).startswith("oc_"):
                ret_send = _send_text(rid, content, receive_id_type="chat_id")
            else:
                ret_send = _send_text(rid, content, receive_id_type="open_id")
            ok = ret_send.get("code") == 0
            if ok:
                success += 1
            logs.append(
                {
                    "record_name": record_name,
                    "open_or_chat_id": rid,
                    "content": content,
                    "ok": ok,
                    "raw": ret_send,
                }
            )
        _write_json(SEND_LOG_FILE, logs)
        _save_sender_last_context(sender_open_id=sender_open_id, recipients=recipients, content=content, source=f"record:{record_name}")
        ret = _reply_chat(chat_id, f"按记录群发完成：{record_name}，成功 {success}/{len(recipients)}")
        return {"ok": True, "mode": "send_by_record", "send_result": ret}

    # ---------- 手动命令：学生画像 ----------
    if text == "自动回复状态":
        auto_rows = _load_auto_reply_settings()
        ai_enabled = bool(SETTINGS.ai_base_url and SETTINGS.ai_api_key and SETTINGS.ai_model)
        reply_now, reason = _should_auto_reply_now()
        reason_map = {
            "busy_mode": "忙碌模式",
            "weekend": "周末",
            "off_hours": "非工作时间",
            "work_hours": "工作时间内",
            "disabled": "已关闭",
        }
        faq_count = len(_load_auto_reply_faq())
        msg = (
            f"自动回复状态：{'开启' if auto_rows.get('enabled', True) else '关闭'}\n"
            f"忙碌模式：{'开启' if auto_rows.get('busy_mode', False) else '关闭'}\n"
            f"工作时间：{auto_rows.get('work_start', '09:00')} - {auto_rows.get('work_end', '18:30')}\n"
            f"周末自动回复：{'开启' if auto_rows.get('weekend_auto_reply', True) else '关闭'}\n"
            f"当前是否触发：{'会' if reply_now else '不会'}（{reason_map.get(reason, reason)}）\n"
            f"常见问答条数：{faq_count}\n"
            f"智能体回退：{'已配置' if ai_enabled else '未配置'}"
        )
        ret = _reply_chat(chat_id, msg)
        return {"ok": True, "mode": "auto_reply_status", "send_result": ret}

    if text == "开启自动回复":
        rows = _load_auto_reply_settings()
        rows["enabled"] = True
        _save_auto_reply_settings(rows)
        ret = _reply_chat(chat_id, "已开启自动回复。周末、下班后或忙碌模式下会优先自动回复常见问题。")
        return {"ok": True, "mode": "enable_auto_reply", "send_result": ret}

    if text == "关闭自动回复":
        rows = _load_auto_reply_settings()
        rows["enabled"] = False
        _save_auto_reply_settings(rows)
        ret = _reply_chat(chat_id, "已关闭自动回复。")
        return {"ok": True, "mode": "disable_auto_reply", "send_result": ret}

    if text == "开启忙碌模式":
        rows = _load_auto_reply_settings()
        rows["busy_mode"] = True
        _save_auto_reply_settings(rows)
        ret = _reply_chat(chat_id, "已开启忙碌模式。学生私聊常见问题会先自动回复。")
        return {"ok": True, "mode": "enable_busy_mode", "send_result": ret}

    if text == "关闭忙碌模式":
        rows = _load_auto_reply_settings()
        rows["busy_mode"] = False
        _save_auto_reply_settings(rows)
        ret = _reply_chat(chat_id, "已关闭忙碌模式。")
        return {"ok": True, "mode": "disable_busy_mode", "send_result": ret}

    if text.startswith("设置自动回复时间 "):
        body = text.replace("设置自动回复时间 ", "", 1).strip()
        seg = body.split()
        if len(seg) != 2 or ":" not in seg[0] or ":" not in seg[1]:
            ret = _reply_chat(chat_id, "格式错误，请使用：设置自动回复时间 09:00 18:30")
            return {"ok": True, "mode": "set_auto_reply_window", "send_result": ret}
        rows = _load_auto_reply_settings()
        rows["work_start"] = seg[0].strip()
        rows["work_end"] = seg[1].strip()
        _save_auto_reply_settings(rows)
        ret = _reply_chat(chat_id, f"已设置自动回复工作时间：{seg[0].strip()} - {seg[1].strip()}")
        return {"ok": True, "mode": "set_auto_reply_window", "send_result": ret}

    text_lower = text.lower()
    if "学生画像" in text or "性格分析" in text or "profile" in text_lower or "persona" in text_lower:
        name = "未命名学员"
        parts = text.split()
        if len(parts) >= 2:
            name = parts[1]
        profile = _analyze_student(name=name, text=text)
        card = _build_profile_card(profile)
        target_teacher = SETTINGS.teacher_open_id.strip()
        if not _is_valid_open_id(target_teacher):
            target_teacher = sender_open_id
        if not target_teacher:
            raise HTTPException(status_code=400, detail="无法确定接收画像卡片的老师 open_id")
        send_ret = _send_card(target_teacher, card, receive_id_type="open_id")
        fallback_ret = None
        if send_ret.get("code") != 0 and chat_id:
            fallback_ret = _reply_chat(chat_id, f"画像已生成，但发送失败：{send_ret.get('msg', 'unknown')}")
        _append_debug(
            {
                "kind": "manual_profile",
                "event_id": event_id,
                "sender_open_id": sender_open_id,
                "teacher_open_id": target_teacher,
                "chat_id": chat_id,
                "text": text,
                "send_ret": send_ret,
                "fallback_ret": fallback_ret,
            }
        )
        return {"ok": True, "mode": "profile", "to": target_teacher, "send_result": send_ret, "fallback_result": fallback_ret}

    # ---------- 自动预警：无需命令 ----------
    # 场景：学生在与机器人会话中发送任意消息，系统自动分析并在达到阈值时私聊老师。
    teacher_oid = SETTINGS.teacher_open_id.strip()
    is_teacher_sender = bool(teacher_oid and sender_open_id == teacher_oid)
    student_name = _student_name_for_sender(sender_open_id) if not is_teacher_sender else ""
    if student_name:
        profile = _analyze_student(name=student_name, text=text)
        if _allow_alert(profile["risk_level"]):
            alert_receivers = _get_alert_receiver_open_ids()
            if not alert_receivers:
                _append_debug(
                    {
                        "kind": "auto_alert_skipped",
                        "event_id": event_id,
                        "reason": "alert_receiver_missing",
                        "sender_open_id": sender_open_id,
                        "student_name": student_name,
                        "text": text,
                    }
                )
                return {"ok": True, "mode": "auto_monitor", "alert_sent": False, "reason": "alert_receiver_missing"}

            card = _build_profile_card(profile)
            send_results: list[dict[str, Any]] = []
            sent_ok = False
            for receiver_open_id in alert_receivers:
                send_ret = _send_card(receiver_open_id, card, receive_id_type="open_id")
                send_results.append({"open_id": receiver_open_id, "result": send_ret})
                if send_ret.get("code") == 0:
                    sent_ok = True
            _append_debug(
                {
                    "kind": "auto_alert",
                    "event_id": event_id,
                    "sender_open_id": sender_open_id,
                    "student_name": student_name,
                    "alert_receivers": alert_receivers,
                    "risk_level": profile["risk_level"],
                    "text": text,
                    "send_results": send_results,
                }
            )
            return {"ok": True, "mode": "auto_monitor", "alert_sent": sent_ok, "send_results": send_results}

        _append_debug(
            {
                "kind": "auto_alert_not_triggered",
                "event_id": event_id,
                "sender_open_id": sender_open_id,
                "student_name": student_name,
                "risk_level": profile["risk_level"],
                "threshold": SETTINGS.alert_level,
            }
        )
        if message.get("chat_type") == "p2p":
            should_reply, reply_reason = _should_auto_reply_now()
            if should_reply and _can_send_auto_reply(sender_open_id):
                auto_reply_text, auto_reply_source = _build_auto_reply(student_name, text)
                send_ret = _reply_chat(chat_id, auto_reply_text)
                _append_debug(
                    {
                        "kind": "auto_reply",
                        "event_id": event_id,
                        "sender_open_id": sender_open_id,
                        "student_name": student_name,
                        "reason": reply_reason,
                        "source": auto_reply_source,
                        "text": text,
                        "send_ret": send_ret,
                    }
                )
                return {
                    "ok": True,
                    "mode": "auto_monitor",
                    "alert_sent": False,
                    "auto_reply_sent": send_ret.get("code") == 0,
                    "auto_reply_source": auto_reply_source,
                }
        return {"ok": True, "mode": "auto_monitor", "alert_sent": False, "reason": "below_threshold"}

    # 未识别命令时，给老师明确反馈，避免“发了没反应”。
    if sender_open_id == SETTINGS.teacher_open_id or message.get("chat_type") == "p2p":
        ret = _reply_chat(chat_id, "未识别命令。发送“帮助”查看可用命令。")
        return {"ok": True, "mode": "unknown_command", "send_result": ret}

    return {"ok": True, "ignored": "no_command_and_unbound_student"}


@app.post("/notify/template/upsert")
def upsert_template(payload: TemplateUpsert) -> dict[str, Any]:
    rows = _read_json(TEMPLATE_FILE, [])
    rows = [r for r in rows if r.get("template_name") != payload.template_name]
    rows.append(payload.model_dump())
    _write_json(TEMPLATE_FILE, rows)
    return {"ok": True, "template_name": payload.template_name, "count": len(payload.recipient_open_ids)}


@app.post("/notify/send")
def send_by_template(payload: SendRequest) -> dict[str, Any]:
    rows = _read_json(TEMPLATE_FILE, [])
    row = next((r for r in rows if r.get("template_name") == payload.template_name), None)
    if not row:
        raise HTTPException(status_code=404, detail="模板不存在")
    recipients = row.get("recipient_open_ids", [])
    if not recipients:
        raise HTTPException(status_code=400, detail="模板没有接收人")

    logs = _read_json(SEND_LOG_FILE, [])
    success = 0
    result_items: list[dict[str, Any]] = []
    for open_id in recipients:
        ret = _send_text(open_id, payload.content)
        ok = bool(ret.get("mock")) or ret.get("code") == 0
        if ok:
            success += 1
        result_items.append({"open_id": open_id, "ok": ok, "raw": ret})
        logs.append({"template_name": payload.template_name, "open_id": open_id, "content": payload.content, "ok": ok, "raw": ret})
    _write_json(SEND_LOG_FILE, logs)
    return {"ok": True, "template_name": payload.template_name, "total": len(recipients), "success": success, "results": result_items}


@app.get("/notify/templates")
def list_templates() -> dict[str, Any]:
    return {"ok": True, "data": _read_json(TEMPLATE_FILE, [])}


@app.get("/notify/logs")
def list_logs() -> dict[str, Any]:
    return {"ok": True, "data": _read_json(SEND_LOG_FILE, [])}
